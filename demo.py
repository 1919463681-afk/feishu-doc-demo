from flask import Flask, request, jsonify
import csv
import difflib
import json
import os
import re
import requests
import threading
import time
import uuid
from datetime import datetime

app = Flask(__name__)

# ==========================================
# 1. 配置
# ==========================================
# 密钥与业务 token：复制 local_config.example.py → local_config.py 填写；
# 或通过环境变量 FEISHU_APP_ID / FEISHU_APP_SECRET 等注入（见 _load_config）。


def _load_config():
    """从 local_config.py 或环境变量加载敏感配置。"""
    lc = {}
    try:
        import local_config as _lc
        lc = {
            k: v for k, v in vars(_lc).items()
            if not k.startswith("_")
        }
    except ImportError:
        pass

    def _str(key, env_key, default=""):
        env_val = os.environ.get(env_key, "").strip()
        if env_val:
            return env_val
        val = lc.get(key, default)
        return val if val is not None else default

    def _list(key, env_key):
        env_val = os.environ.get(env_key, "").strip()
        if env_val:
            return [x.strip() for x in env_val.split(",") if x.strip()]
        val = lc.get(key)
        return list(val) if val else []

    return {
        "APP_ID": _str("APP_ID", "FEISHU_APP_ID"),
        "APP_SECRET": _str("APP_SECRET", "FEISHU_APP_SECRET"),
        "AUDIT_ROOT_FOLDER_TOKEN": _str("AUDIT_ROOT_FOLDER_TOKEN", "FEISHU_AUDIT_FOLDER_TOKEN"),
        "AUDIT_WIKI_SEED_NODE": _str("AUDIT_WIKI_SEED_NODE", "FEISHU_WIKI_SEED_NODE"),
        "AUDIT_WIKI_SPACE_IDS": _list("AUDIT_WIKI_SPACE_IDS", "FEISHU_WIKI_SPACE_IDS"),
        "AUDIT_DOC_TOKENS": _list("AUDIT_DOC_TOKENS", "FEISHU_AUDIT_DOC_TOKENS"),
    }


_cfg = _load_config()
APP_ID = _cfg["APP_ID"]
APP_SECRET = _cfg["APP_SECRET"]
AUDIT_ROOT_FOLDER_TOKEN = _cfg["AUDIT_ROOT_FOLDER_TOKEN"]
AUDIT_WIKI_SPACE_IDS = _cfg["AUDIT_WIKI_SPACE_IDS"]
AUDIT_WIKI_SEED_NODE = _cfg["AUDIT_WIKI_SEED_NODE"]
AUDIT_DOC_TOKENS = _cfg["AUDIT_DOC_TOKENS"]

if not APP_ID or not APP_SECRET:
    raise SystemExit(
        "缺少飞书应用凭证：请复制 local_config.example.py 为 local_config.py 并填写 APP_ID/APP_SECRET，"
        "或设置环境变量 FEISHU_APP_ID、FEISHU_APP_SECRET"
    )

# 敏感词库（demo 示例，可按需扩展或接入大模型）
SENSITIVE_WORDS = ["违禁词", "测试敏感", "暴力", "赌博"]

# 手动配置要全量审核的文档 token（应用已添加为「文档应用」的文档）
# 见上方 AUDIT_DOC_TOKENS（local_config.py / 环境变量）

# 自动记录 subscribe / webhook 触达过的文档
KNOWN_DOCS_FILE = os.path.join(os.path.dirname(__file__), "known_docs.json")
AUDIT_REPORT_DIR = os.path.join(os.path.dirname(__file__), "audit_reports")
EVENTS_LOG_FILE = os.path.join(os.path.dirname(__file__), "document_events.jsonl")
EVENTS_LOG_MAX = 2000

# 用户行为采集表（对照需求）—— 快照与活动类型
CONTENT_SNAPSHOT_FILE = os.path.join(os.path.dirname(__file__), "content_snapshot.json")
PUBLIC_PERM_SNAPSHOT_FILE = os.path.join(os.path.dirname(__file__), "public_permission_snapshot.json")
COMMENT_SNAPSHOT_FILE = os.path.join(os.path.dirname(__file__), "comment_snapshot.json")
ACTIVITY_TYPES = {
    "create": "创建",
    "view": "查看",
    "edit": "编辑",
    "comment": "评论",
    "download": "下载",
    "forward": "转发",
    "share": "分享",
    "delete": "删除",
}
LINK_SHARE_LABELS = {
    "tenant_readable": "组织内链接只读",
    "tenant_editable": "组织内链接可编辑",
    "anyone_readable": "互联网链接只读(公开)",
    "anyone_editable": "互联网链接可编辑(公开)",
    "closed": "关闭链接分享",
}
PERM_LABELS = {
    "view": "只读",
    "edit": "可编辑",
    "full_access": "可管理",
}

# 监听的云文档事件（需在飞书开放平台「事件订阅」中一并勾选）
MONITORED_EVENT_TYPES = {
    "drive.file.created_in_folder_v1": "新建文件",
    "drive.file.edit_v1": "内容编辑",
    "drive.file.permission_member_added_v1": "权限-添加协作者",
    "drive.file.permission_member_removed_v1": "权限-移除协作者",
    "drive.file.permission_member_applied_v1": "权限-申请协作者",
    "drive.file.trashed_v1": "删除到回收站",
    "drive.file.deleted_v1": "彻底删除",
    "drive.notice.comment_add_v1": "评论/回复",
    # 方案 B：轮询发现（应用非所有者时替代 webhook）
    "poll.permission_member_added": "权限-添加协作者(轮询)",
    "poll.permission_member_removed": "权限-移除协作者(轮询)",
    "poll.permission_member_changed": "权限-变更(轮询)",
    "poll.file_deleted": "文件已删除(轮询)",
    "poll.public_permission_changed": "分享设置变更(轮询)",
    "poll.comment_added": "评论(轮询)",
    "track.view": "查看(埋点)",
    "track.download": "下载(埋点)",
}
# 飞书限制：应用仅为文档「可管理」协作者时，实际只能收到以下事件（收不到权限/删除）
# 参考 https://open.feishu.cn/document/server-docs/docs/drive-v1/event/subscribe
MANAGER_RECEIVABLE_EVENTS = frozenset({
    "drive.file.edit_v1",
    "drive.file.bitable_field_changed_v1",
    "drive.file.bitable_record_changed_v1",
})
# 开放平台「事件订阅」需添加的事件 + 建议权限（诊断用）
FEISHU_EVENT_SETUP_CHECKLIST = [
    ("drive.file.created_in_folder_v1", "space:document.event:read（应用身份+用户身份）"),
    ("drive.file.edit_v1", "docs:event:subscribe"),
    ("drive.file.permission_member_added_v1", "docs:permission.member:retrieve"),
    ("drive.file.permission_member_removed_v1", "docs:permission.member:retrieve"),
    ("drive.file.permission_member_applied_v1", "docs:permission.member:create 或 retrieve"),
    ("drive.file.trashed_v1", "docs:event.document_deleted:read"),
    ("drive.file.deleted_v1", "docs:event.document_deleted:read"),
    ("drive.notice.comment_add_v1", "docs:document.comment:read"),
]
# 评论 Webhook 额外配置说明（需在开放平台「事件订阅」添加，与 drive.file.* 不同）
FEISHU_COMMENT_WEBHOOK_SETUP = [
    "开放平台 → 事件订阅 → 添加事件 drive.notice.comment_add_v1",
    "权限：docs:document.comment:read（获取云文档中的评论）",
    "Request URL 与编辑事件相同，指向 /webhook（需 cpolar 等内网穿透）",
    "启动时自动调用 POST /drive/v1/user/subscription 订阅评论通知（本服务已实现）",
    "评论是「通知类」事件：仅当飞书会给某人发评论通知时才推送（自己评自己/无@可能无推送）",
    "建议测试：用另一个账号评论，或评论时 @ 文档协作者",
    "局部(划词)评论正文通过 batch_query + replies API 补全，非事件体直接携带",
]
# 收到以下事件时自动做内容审核
CONTENT_AUDIT_ON_EVENTS = frozenset({
    "drive.file.created_in_folder_v1",
    "drive.file.edit_v1",
})
# 启动 Flask 时自动订阅共享文件夹「新建文件」事件（需配置 AUDIT_ROOT_FOLDER_TOKEN）
AUTO_SUBSCRIBE_FOLDER_ON_START = True
# 启动时自动 subscribe 共享文件夹内已有文件（权限/删除事件的前提）
AUTO_SUBSCRIBE_ALL_FILES_ON_START = True
# 收到新建事件后，自动 subscribe 该文件（后续才能收到编辑/权限变更）
AUTO_SUBSCRIBE_NEW_FILE = True
AUTO_SUBSCRIBE_USER_COMMENT_ON_START = True
WEBHOOK_DEBUG_LOG = os.path.join(os.path.dirname(__file__), "webhook_debug.jsonl")

# 方案 B：轮询监听（权限变更 + 文件删除，不依赖文档所有者 webhook）
PERMISSION_SNAPSHOT_FILE = os.path.join(os.path.dirname(__file__), "permission_snapshot.json")
FOLDER_FILES_SNAPSHOT_FILE = os.path.join(os.path.dirname(__file__), "folder_files_snapshot.json")
POLL_INTERVAL_SECONDS = 300
AUTO_POLL_ON_START = True
POLLABLE_PERMISSION_TYPES = frozenset({"docx", "doc", "sheet", "bitable", "slides", "file"})

# 权限审计导出里常见的 token / 标题列名
TOKEN_COLUMN_HINTS = (
    "token", "doc_token", "docs_token", "file_token", "document_id",
    "云文档id", "云文档 token", "文档id", "文档 token", "文件id", "对象id",
)
TITLE_COLUMN_HINTS = ("title", "标题", "文档名称", "文档标题", "名称", "文件名")
TYPE_COLUMN_HINTS = ("type", "docs_type", "类型", "文档类型", "文件类型")
DOCX_TYPE_VALUES = {"docx", "doc", "文档", "新版文档", "旧版文档"}

# 云盘 / wiki 全量扫描时纳入的文件类型（不含 folder、shortcut）
AUDIT_FILE_TYPES = frozenset({
    "docx", "doc", "sheet", "bitable", "slides", "mindnote",
})
# 可通过开放 API 读取正文并做敏感词审核的类型
AUDITABLE_CONTENT_TYPES = frozenset({"docx", "doc", "sheet", "bitable", "slides"})
# 飞书 API 暂不支持读取正文的类型（仍会出现在扫描列表里并标注跳过原因）
UNSUPPORTED_CONTENT_TYPES = frozenset({"mindnote"})


# ==========================================
# 鉴权
# ==========================================
def get_tenant_access_token():
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    headers = {"Content-Type": "application/json; charset=utf-8"}
    payload = {"app_id": APP_ID, "app_secret": APP_SECRET}
    try:
        res = requests.post(url, json=payload, headers=headers)
        return res.json().get("tenant_access_token")
    except Exception as e:
        print(f"获取 Token 失败: {e}")
        return None


def _auth_headers():
    token = get_tenant_access_token()
    if not token:
        return None
    return {"Authorization": f"Bearer {token}"}


# ==========================================
# 文档订阅
# ==========================================
def subscribe_document(file_token, file_type="docx"):
    headers = _auth_headers()
    if not headers:
        return {"code": -1, "msg": "无法获取 tenant_access_token"}

    url = f"https://open.feishu.cn/open-apis/drive/v1/files/{file_token}/subscribe"
    res = requests.post(url, headers=headers, params={"file_type": file_type})
    return res.json()


def subscribe_folder(folder_token, event_type="file.created_in_folder_v1"):
    """订阅文件夹内新建文件事件（用于实时发现新文档）"""
    headers = _auth_headers()
    if not headers:
        return {"code": -1, "msg": "无法获取 tenant_access_token"}

    url = f"https://open.feishu.cn/open-apis/drive/v1/files/{folder_token}/subscribe"
    res = requests.post(
        url,
        headers=headers,
        params={"file_type": "folder", "event_type": event_type},
    )
    return res.json()


def subscribe_user_comment_notice():
    """
    订阅用户云文档评论/回复通知（drive.notice.comment_add_v1 的 Webhook 前提之一）。
    与 drive.file.edit_v1 不同，评论是「通知类」事件，需单独调用此接口。
    """
    headers = _auth_headers()
    if not headers:
        return {"code": -1, "msg": "无法获取 tenant_access_token"}

    url = "https://open.feishu.cn/open-apis/drive/v1/user/subscription"
    res = requests.post(
        url,
        headers=headers,
        json={"event_type": "drive.notice.comment_add_v1"},
        timeout=15,
    )
    return res.json()


def auto_subscribe_new_file(file_token, file_type, source="auto_new", title=None):
    """
    新建文件后自动 subscribe，使该文件后续能收到编辑/权限变更事件。
    返回 subscribe API 的响应 dict。
    """
    if not AUTO_SUBSCRIBE_NEW_FILE:
        return {"code": -1, "msg": "AUTO_SUBSCRIBE_NEW_FILE 已关闭", "skipped": True}

    raw_type = (file_type or "docx").lower()
    if raw_type in UNSUPPORTED_CONTENT_TYPES:
        return {"code": -1, "msg": f"类型 {raw_type} 不支持 subscribe", "skipped": True}

    ftype = normalize_file_type(raw_type)
    result = subscribe_document(file_token, ftype)
    if result.get("code") == 0:
        register_known_doc(file_token, title=title, source=source, file_type=ftype)
        print(f"   ✅ 新建自动订阅成功: {file_token} ({ftype})")
    else:
        print(f"   ⚠️ 新建自动订阅失败: {file_token} ({ftype}) -> {result.get('msg', result)}")
    return result


def ensure_folder_subscribe_on_start():
    """服务启动时订阅共享文件夹 + 已有文件，确保能收到新建/编辑/权限/删除事件"""
    if not AUDIT_ROOT_FOLDER_TOKEN:
        return None

    folder_result = None
    if AUTO_SUBSCRIBE_FOLDER_ON_START:
        folder_result = subscribe_folder(AUDIT_ROOT_FOLDER_TOKEN)
        if folder_result.get("code") == 0:
            print(f"🔄 已自动订阅文件夹新建事件: {AUDIT_ROOT_FOLDER_TOKEN}")
        else:
            print(f"⚠️ 文件夹新建事件订阅失败: {folder_result.get('msg', folder_result)}")

    if AUTO_SUBSCRIBE_ALL_FILES_ON_START:
        ok, fail = 0, 0
        for f in collect_all_folder_files(AUDIT_ROOT_FOLDER_TOKEN):
            token = f.get("token")
            raw_type = (f.get("type") or "docx").lower()
            if raw_type in UNSUPPORTED_CONTENT_TYPES or raw_type not in AUDIT_FILE_TYPES:
                continue
            ftype = normalize_file_type(raw_type)
            sub = subscribe_document(token, ftype)
            if sub.get("code") == 0:
                ok += 1
                register_known_doc(token, title=f.get("name"), source="auto_start", file_type=ftype)
            else:
                fail += 1
                print(f"   ⚠️ 启动订阅失败 {f.get('name')}: {sub.get('msg', sub)}")
        print(f"🔄 启动时已订阅文件夹内 {ok} 个文件" + (f"，失败 {fail} 个" if fail else ""))

    if AUTO_SUBSCRIBE_USER_COMMENT_ON_START:
        sub = subscribe_user_comment_notice()
        if sub.get("code") == 0:
            print("🔄 已订阅用户云文档评论通知（drive.notice.comment_add_v1）")
        else:
            print(f"⚠️ 用户评论通知订阅失败: {sub.get('msg', sub)}")

    return folder_result


def check_subscribe_status(file_token, file_type="docx"):
    """查询单文件 subscribe 状态"""
    status = get_document_subscribe_status(file_token, file_type)
    subscribed = status.get("data", {}).get("is_subscribe")
    return {
        "file_token": file_token,
        "file_type": file_type,
        "is_subscribe": subscribed,
        "api_response": status,
        "hint": (
            "已订阅；权限/删除事件还需：开放平台勾选对应事件 + 应用为文档所有者"
            if subscribed else
            "未订阅；请运行 subscribe 或 subscribe-all"
        ),
    }


def diagnose_event_monitoring(folder_token=None):
    """诊断事件监听配置（subscribe 状态 + 检查清单）"""
    folder = folder_token or AUDIT_ROOT_FOLDER_TOKEN
    files = collect_all_folder_files(folder or "")
    file_checks = []
    for f in files[:20]:
        token = f.get("token")
        raw_type = (f.get("type") or "docx").lower()
        if raw_type in UNSUPPORTED_CONTENT_TYPES:
            continue
        ftype = normalize_file_type(raw_type)
        file_checks.append({
            "title": f.get("name"),
            **check_subscribe_status(token, ftype),
        })

    unsubscribed = [x for x in file_checks if not x.get("is_subscribe")]
    return {
        "folder_token": folder,
        "files_checked": len(file_checks),
        "unsubscribed_count": len(unsubscribed),
        "file_checks": file_checks,
        "feishu_console_checklist": [
            {"event": e, "permission": p} for e, p in FEISHU_EVENT_SETUP_CHECKLIST
        ],
        "feishu_comment_webhook_setup": FEISHU_COMMENT_WEBHOOK_SETUP,
        "important_limits": [
            "权限变更、删除事件：应用必须是文档「所有者」，仅「可管理」协作者收不到（飞书 API 限制）",
            "文档应用添加方式：文档右上角 … → 更多 → 添加文档应用 → 选「可管理」；若仍收不到权限/删除，需将文档所有权转给应用或改用用户身份订阅",
            "开放平台「事件订阅」必须勾选 permission_member_* / trashed / deleted / drive.notice.comment_add_v1 并重新发版",
            "评论实时事件：drive.notice.comment_add_v1 + docs:document.comment:read，Webhook 地址同 /webhook",
            "局部评论正文：事件不含全文，需 batch_query + replies API 补全（本服务已自动调用）",
            "删除事件需额外权限 docs:event.document_deleted:read",
            "编辑能收到但权限/删除收不到 → 典型原因是应用只是管理者而非所有者",
        ],
    }


def get_document_subscribe_status(file_token, file_type="docx"):
    headers = _auth_headers()
    if not headers:
        return {"code": -1, "msg": "无法获取 tenant_access_token"}

    url = f"https://open.feishu.cn/open-apis/drive/v1/files/{file_token}/get_subscribe"
    res = requests.get(url, headers=headers, params={"file_type": file_type})
    return res.json()


def subscribe_all_monitors(folder_token=None):
    """
    一键订阅实时监听：
    1. 共享文件夹「新建文件」事件
    2. 文件夹内已有文件的全部云文档事件（编辑、权限变更等）
    3. wiki 知识库内已扫描到的页面
    """
    folder = folder_token if folder_token is not None else AUDIT_ROOT_FOLDER_TOKEN
    summary = {
        "folder_token": folder,
        "folder_subscribe": None,
        "file_subscribes": [],
        "wiki_subscribes": [],
        "success": 0,
        "failed": 0,
    }

    if folder:
        summary["folder_subscribe"] = subscribe_folder(folder)
        if summary["folder_subscribe"].get("code") == 0:
            summary["success"] += 1
        else:
            summary["failed"] += 1
        print(f"📁 文件夹新建事件订阅: {summary['folder_subscribe']}")

    for f in collect_all_folder_files(folder or ""):
        token = f.get("token")
        raw_type = (f.get("type") or "docx").lower()
        if raw_type not in AUDIT_FILE_TYPES or raw_type in UNSUPPORTED_CONTENT_TYPES:
            continue
        ftype = normalize_file_type(raw_type)
        sub = subscribe_document(token, ftype)
        entry = {
            "token": token,
            "file_type": ftype,
            "title": f.get("name"),
            "result": sub,
        }
        summary["file_subscribes"].append(entry)
        if sub.get("code") == 0:
            summary["success"] += 1
            register_known_doc(token, title=f.get("name"), source="subscribe_all", file_type=ftype)
        else:
            summary["failed"] += 1
        print(f"   📄 {f.get('name')} ({ftype}): code={sub.get('code')}")

    for space_id in collect_wiki_space_ids():
        print(f"📚 wiki 订阅 space_id={space_id}")
        for item in collect_wiki_file_nodes(space_id):
            node_token = item.get("node_token")
            if not node_token:
                continue
            sub = subscribe_wiki_node(node_token)
            entry = {"node_token": node_token, "title": item.get("title"), "result": sub}
            summary["wiki_subscribes"].append(entry)
            if sub.get("code") == 0:
                summary["success"] += 1
                register_known_doc(
                    item.get("token"),
                    title=item.get("title"),
                    source="subscribe_all_wiki",
                    file_type=item.get("file_type"),
                )
            else:
                summary["failed"] += 1

    print(f"✅ 订阅完成: 成功 {summary['success']}，失败 {summary['failed']}")
    return summary


# ==========================================
# 内容审核核心
# ==========================================
def fetch_document_raw_content(document_id):
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    url = f"https://open.feishu.cn/open-apis/docx/v1/documents/{document_id}/raw_content"
    res = requests.get(url, headers=headers)
    data = res.json()
    if data.get("code") != 0:
        return None, data.get("msg", "拉取文档失败")
    return data.get("data", {}).get("content", ""), None


def fetch_legacy_doc_content(doc_token):
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    url = f"https://open.feishu.cn/open-apis/doc/v2/{doc_token}/raw_content"
    res = requests.get(url, headers=headers)
    data = res.json()
    if data.get("code") != 0:
        return None, data.get("msg", "拉取旧版文档失败")
    return data.get("data", {}).get("content", ""), None


def _flatten_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return " ".join(_flatten_cell_value(v) for v in value)
    if isinstance(value, dict):
        for key in ("text", "name", "link", "url", "value"):
            if key in value and value[key]:
                return _flatten_cell_value(value[key])
        return " ".join(_flatten_cell_value(v) for v in value.values() if v is not None)
    return str(value)


def fetch_sheet_content(spreadsheet_token):
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    meta_url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo"
    meta_res = requests.get(meta_url, headers=headers)
    meta_data = meta_res.json()
    if meta_data.get("code") != 0:
        return None, meta_data.get("msg", "拉取表格元数据失败")

    sheets = meta_data.get("data", {}).get("sheets", [])
    if not sheets:
        return "", None

    parts = []
    for sheet in sheets:
        sheet_id = sheet.get("sheetId")
        if not sheet_id:
            continue
        title = sheet.get("title") or sheet_id
        row_count = min(int(sheet.get("rowCount") or 500), 500)
        range_str = f"{sheet_id}!A1:ZZ{row_count}"
        val_url = (
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
            f"{spreadsheet_token}/values/{range_str}"
        )
        val_res = requests.get(
            val_url,
            headers=headers,
            params={"valueRenderOption": "ToString"},
        )
        val_data = val_res.json()
        if val_data.get("code") != 0:
            parts.append(f"[{title}] 读取失败: {val_data.get('msg', '')}")
            continue
        rows = val_data.get("data", {}).get("valueRange", {}).get("values") or []
        row_texts = []
        for row in rows:
            cells = [_flatten_cell_value(c) for c in row if c not in (None, "")]
            if cells:
                row_texts.append(" ".join(cells))
        if row_texts:
            parts.append(f"[{title}]\n" + "\n".join(row_texts))

    return "\n\n".join(parts), None


def fetch_bitable_content(app_token):
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    tables_url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables"
    tables_res = requests.get(tables_url, headers=headers, params={"page_size": 100})
    tables_data = tables_res.json()
    if tables_data.get("code") != 0:
        return None, tables_data.get("msg", "拉取多维表格失败")

    tables = tables_data.get("data", {}).get("items", [])
    parts = []
    for table in tables:
        table_id = table.get("table_id")
        table_name = table.get("name") or table_id
        if not table_id:
            continue

        search_url = (
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}"
            f"/tables/{table_id}/records/search"
        )
        page_token = None
        record_texts = []
        while True:
            body = {"page_size": 500}
            if page_token:
                body["page_token"] = page_token
            search_res = requests.post(search_url, headers=headers, json=body)
            search_data = search_res.json()
            if search_data.get("code") != 0:
                record_texts.append(f"读取失败: {search_data.get('msg', '')}")
                break

            payload = search_data.get("data", {})
            for record in payload.get("items", []):
                fields = record.get("fields") or {}
                cell_texts = [_flatten_cell_value(v) for v in fields.values()]
                cell_texts = [t for t in cell_texts if t]
                if cell_texts:
                    record_texts.append(" ".join(cell_texts))

            if not payload.get("has_more"):
                break
            page_token = payload.get("page_token")
            if not page_token:
                break

        if record_texts:
            parts.append(f"[{table_name}]\n" + "\n".join(record_texts))

    return "\n\n".join(parts), None


def fetch_slides_content(presentation_token):
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    list_url = (
        f"https://open.feishu.cn/open-apis/slides/v1/presentations/"
        f"{presentation_token}/slides"
    )
    list_res = requests.get(list_url, headers=headers)
    try:
        list_data = list_res.json()
    except ValueError:
        return None, f"幻灯片 API 返回非 JSON（HTTP {list_res.status_code}），请确认已开通 slides 相关权限"

    if list_data.get("code") != 0:
        return None, list_data.get("msg", "拉取幻灯片失败")

    slides = list_data.get("data", {}).get("slides", [])
    parts = []
    for slide in slides:
        slide_id = slide.get("slide_id")
        if not slide_id:
            continue
        detail_url = (
            f"https://open.feishu.cn/open-apis/slides/v1/presentations/"
            f"{presentation_token}/slides/{slide_id}"
        )
        detail_res = requests.get(detail_url, headers=headers)
        try:
            detail_data = detail_res.json()
        except ValueError:
            continue
        if detail_data.get("code") != 0:
            continue
        slide_obj = detail_data.get("data", {}).get("slide", {})
        texts = []

        def _walk(obj):
            if isinstance(obj, dict):
                for key in ("text", "content", "title", "body"):
                    if key in obj and obj[key]:
                        texts.append(_flatten_cell_value(obj[key]))
                for v in obj.values():
                    _walk(v)
            elif isinstance(obj, list):
                for item in obj:
                    _walk(item)

        _walk(slide_obj)
        if texts:
            parts.append("\n".join(texts))

    if parts:
        return "\n\n".join(parts), None
    return "", None


def fetch_file_content(file_token, file_type):
    """按文件类型拉取可审核的正文"""
    ftype = (file_type or "docx").lower()
    if ftype in UNSUPPORTED_CONTENT_TYPES:
        return None, f"飞书暂未开放 {ftype} 类型的正文读取 API，无法自动审核"
    if ftype == "doc":
        return fetch_legacy_doc_content(file_token)
    if ftype not in AUDITABLE_CONTENT_TYPES:
        return None, f"不支持的文件类型: {ftype}"

    if ftype == "docx":
        return fetch_document_raw_content(file_token)
    if ftype == "sheet":
        return fetch_sheet_content(file_token)
    if ftype == "bitable":
        return fetch_bitable_content(file_token)
    if ftype == "slides":
        return fetch_slides_content(file_token)
    return None, f"未实现的类型: {ftype}"


def normalize_file_type(file_type):
    ftype = (file_type or "docx").lower()
    if ftype == "doc":
        return "docx"
    return ftype


def moderate_text(text):
    if not text or not text.strip():
        return {"passed": True, "hits": [], "summary": "文档为空"}

    hits = [word for word in SENSITIVE_WORDS if word in text]
    passed = len(hits) == 0
    summary = "审核通过" if passed else f"发现 {len(hits)} 处敏感词"
    return {"passed": passed, "hits": hits, "summary": summary}


def audit_file(file_token, file_type="docx", source="manual", title=None):
    """审核单篇云文档（支持 docx / sheet / bitable / slides 等）"""
    ftype = normalize_file_type(file_type)
    label = f"{file_token} ({ftype})"
    print(f"🔍 [{source}] 开始审核: {label}" + (f" {title}" if title else ""))

    content, err = fetch_file_content(file_token, file_type)
    if err:
        result = {
            "file_token": file_token,
            "file_type": ftype,
            "document_id": file_token,
            "title": title,
            "source": source,
            "passed": None,
            "error": err,
        }
        print(f"❌ [{source}] 审核失败 {label}: {err}")
        return result

    moderation = moderate_text(content)
    preview = content[:120].replace("\n", " ") + ("..." if len(content) > 120 else "")
    result = {
        "file_token": file_token,
        "file_type": ftype,
        "document_id": file_token,
        "title": title,
        "source": source,
        "passed": moderation["passed"],
        "hits": moderation["hits"],
        "summary": moderation["summary"],
        "preview": preview,
        "char_count": len(content),
    }

    if moderation["passed"]:
        print(f"✅ [{source}] {label} 审核通过，共 {len(content)} 字")
    else:
        print(f"⚠️ [{source}] {label} 未通过，命中: {moderation['hits']}")

    return result


def audit_document(document_id, source="manual", title=None, file_type="docx"):
    """兼容旧接口：默认审核 docx"""
    return audit_file(document_id, file_type=file_type, source=source, title=title)


def list_files_in_folder(folder_token="", page_token=None):
    headers = _auth_headers()
    if not headers:
        return [], False, None, "无法获取 token"

    params = {"page_size": 200}
    if folder_token:
        params["folder_token"] = folder_token
    if page_token:
        params["page_token"] = page_token

    url = "https://open.feishu.cn/open-apis/drive/v1/files"
    res = requests.get(url, headers=headers, params=params)
    data = res.json()
    if data.get("code") != 0:
        return [], False, None, data.get("msg", "列举文件失败")

    payload = data.get("data", {})
    return payload.get("files", []), payload.get("has_more", False), payload.get("page_token"), None


def load_known_docs():
    if not os.path.exists(KNOWN_DOCS_FILE):
        return {}
    try:
        with open(KNOWN_DOCS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def save_known_docs(docs):
    with open(KNOWN_DOCS_FILE, "w", encoding="utf-8") as f:
        json.dump(docs, f, ensure_ascii=False, indent=2)


def register_known_doc(file_token, title=None, source="manual", file_type=None,
                       folder_token=None, wiki_space_id=None, wiki_space_name=None):
    if not file_token:
        return
    docs = load_known_docs()
    entry = docs.get(file_token, {})
    entry["title"] = title or entry.get("title")
    entry["source"] = source
    if file_type:
        entry["file_type"] = normalize_file_type(file_type)
    if folder_token:
        entry["folder_token"] = folder_token
    if wiki_space_id:
        entry["wiki_space_id"] = wiki_space_id
    if wiki_space_name:
        entry["wiki_space_name"] = wiki_space_name
    docs[file_token] = entry
    save_known_docs(docs)


def get_wiki_node(node_token):
    """wiki URL 里 /wiki/ 后面是 node_token，需转成 obj_token 才能读正文"""
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    url = "https://open.feishu.cn/open-apis/wiki/v2/spaces/get_node"
    res = requests.get(url, headers=headers, params={"token": node_token})
    data = res.json()
    if data.get("code") != 0:
        return None, data.get("msg", "获取 wiki 节点失败")
    return data.get("data", {}).get("node"), None


def resolve_wiki_space_id(node_token):
    node, err = get_wiki_node(node_token)
    if err:
        return None, err
    return node.get("space_id"), None


def list_wiki_child_nodes(space_id, parent_node_token=None, page_token=None):
    headers = _auth_headers()
    if not headers:
        return [], False, None, "无法获取 token"

    params = {"page_size": 50}
    if parent_node_token:
        params["parent_node_token"] = parent_node_token
    if page_token:
        params["page_token"] = page_token

    url = f"https://open.feishu.cn/open-apis/wiki/v2/spaces/{space_id}/nodes"
    res = requests.get(url, headers=headers, params=params)
    data = res.json()
    if data.get("code") != 0:
        return [], False, None, data.get("msg", "列举 wiki 节点失败")

    payload = data.get("data", {})
    return payload.get("items", []), payload.get("has_more", False), payload.get("page_token"), None


def collect_wiki_file_nodes(space_id):
    """递归收集知识库内所有可扫描类型的节点"""
    file_nodes = []
    queue = [None]

    while queue:
        parent = queue.pop(0)
        page_token = None

        while True:
            items, has_more, page_token, err = list_wiki_child_nodes(space_id, parent, page_token)
            if err:
                print(f"⚠️ wiki 列举失败 space={space_id} parent={parent}: {err}")
                break

            for item in items:
                obj_type = (item.get("obj_type") or "").lower()
                if obj_type in AUDIT_FILE_TYPES:
                    file_nodes.append({
                        "token": item.get("obj_token"),
                        "file_type": obj_type,
                        "title": item.get("title"),
                        "node_token": item.get("node_token"),
                        "from": "wiki_scan",
                        "space_id": space_id,
                    })
                if item.get("has_child"):
                    queue.append(item.get("node_token"))

            if not has_more:
                break

    return file_nodes


def collect_wiki_space_ids():
    ids = list(AUDIT_WIKI_SPACE_IDS)
    if AUDIT_WIKI_SEED_NODE and AUDIT_WIKI_SEED_NODE not in ids:
        space_id, err = resolve_wiki_space_id(AUDIT_WIKI_SEED_NODE)
        if space_id and space_id not in ids:
            ids.append(space_id)
            print(f"ℹ️  从 wiki 节点 {AUDIT_WIKI_SEED_NODE} 解析到 space_id: {space_id}")
        elif err:
            print(f"⚠️ 无法从 AUDIT_WIKI_SEED_NODE 解析 space_id: {err}")
    return ids


def audit_wiki_node(node_token):
    """审核单篇 wiki 页面（自动 node_token → obj_token）"""
    node, err = get_wiki_node(node_token)
    if err:
        return {"node_token": node_token, "passed": None, "error": err}

    obj_type = (node.get("obj_type") or "").lower()
    obj_token = node.get("obj_token")
    title = node.get("title")

    if obj_type not in AUDIT_FILE_TYPES:
        return {
            "node_token": node_token,
            "obj_type": obj_type,
            "passed": None,
            "error": f"不在扫描范围内的类型: {obj_type}",
        }

    register_known_doc(
        obj_token, title=title, source="wiki", file_type=obj_type,
        wiki_space_id=node.get("space_id"),
    )
    result = audit_file(obj_token, file_type=obj_type, source="wiki", title=title)
    result["node_token"] = node_token
    return result


def subscribe_wiki_node(node_token):
    """wiki 文档编辑事件走底层云文档的 subscribe"""
    node, err = get_wiki_node(node_token)
    if err:
        return {"code": -1, "msg": err}
    obj_type = normalize_file_type(node.get("obj_type", "docx"))
    return subscribe_document(node.get("obj_token"), obj_type)


def collect_all_folder_files(folder_token=""):
    """递归收集共享文件夹内所有可扫描类型的云文档"""
    collected = []
    folder_queue = [folder_token]
    type_counts = {}

    while folder_queue:
        current_folder = folder_queue.pop(0)
        page_token = None

        while True:
            files, has_more, page_token, err = list_files_in_folder(current_folder, page_token)
            if err:
                print(f"⚠️ 列举文件夹失败 ({current_folder or 'root'}): {err}")
                break

            for f in files:
                ftype = (f.get("type") or "").lower()
                if ftype == "folder":
                    folder_queue.append(f.get("token"))
                elif ftype in AUDIT_FILE_TYPES:
                    collected.append(f)
                    type_counts[ftype] = type_counts.get(ftype, 0) + 1

            if not has_more:
                break

    if type_counts:
        summary = ", ".join(f"{k}={v}" for k, v in sorted(type_counts.items()))
        print(f"📁 文件夹扫描汇总: {summary}")
    return collected


def collect_docs_for_audit(folder_token=None):
    """
    汇总待审核文档，来源：
    1. 云盘 folder 扫描（docx / sheet / bitable / slides / mindnote 等）
    2. 知识库 wiki 空间扫描
    3. AUDIT_DOC_TOKENS / known_docs.json
    """
    root = folder_token if folder_token is not None else AUDIT_ROOT_FOLDER_TOKEN
    merged = {}

    folder_files = collect_all_folder_files(root)
    for f in folder_files:
        token = f.get("token")
        ftype = (f.get("type") or "docx").lower()
        if token:
            merged[token] = {
                "title": f.get("name"),
                "file_type": ftype,
                "from": "folder_scan",
            }

    for space_id in collect_wiki_space_ids():
        print(f"📚 扫描知识库 space_id={space_id}")
        for item in collect_wiki_file_nodes(space_id):
            token = item.get("token")
            if token:
                merged[token] = {
                    "title": item.get("title"),
                    "file_type": item.get("file_type", "docx"),
                    "from": "wiki_scan",
                    "node_token": item.get("node_token"),
                    "wiki_space_id": item.get("space_id") or space_id,
                }
                register_known_doc(
                    token, title=item.get("title"), source="wiki_scan",
                    file_type=item.get("file_type"), wiki_space_id=item.get("space_id") or space_id,
                )

    for token in AUDIT_DOC_TOKENS:
        merged[token] = merged.get(token, {"title": None, "file_type": "docx", "from": "config"})

    for token, meta in load_known_docs().items():
        entry = merged.get(token, {
            "title": meta.get("title"),
            "file_type": meta.get("file_type", "docx"),
            "from": "registry",
        })
        if not entry.get("title"):
            entry["title"] = meta.get("title")
        if not entry.get("file_type"):
            entry["file_type"] = meta.get("file_type", "docx")
        merged[token] = entry

    return merged


def collect_docs_for_poll(folder_token=None):
    """轮询用：仅共享文件夹 + wiki，不含 known_docs 里可能已删除的 token"""
    root = folder_token if folder_token is not None else AUDIT_ROOT_FOLDER_TOKEN
    merged = {}

    for f in collect_all_folder_files(root):
        token = f.get("token")
        ftype = (f.get("type") or "docx").lower()
        if token:
            merged[token] = {
                "title": f.get("name"),
                "file_type": ftype,
                "from": "folder_scan",
            }

    for space_id in collect_wiki_space_ids():
        print(f"📚 扫描知识库 space_id={space_id}")
        for item in collect_wiki_file_nodes(space_id):
            token = item.get("token")
            if token:
                merged[token] = {
                    "title": item.get("title"),
                    "file_type": item.get("file_type", "docx"),
                    "from": "wiki_scan",
                    "node_token": item.get("node_token"),
                    "wiki_space_id": item.get("space_id") or space_id,
                }

    return merged


def audit_all_documents(folder_token=None):
    """批量审核：云盘 folder + 知识库 wiki + 配置/registry"""
    root = folder_token if folder_token is not None else AUDIT_ROOT_FOLDER_TOKEN
    print(f"📂 云盘 folder token: {root or '(未配置)'}")
    print(f"📚 wiki space 数量: {len(collect_wiki_space_ids())}")
    print(f"ℹ️  扫描类型: {', '.join(sorted(AUDIT_FILE_TYPES))}")

    doc_map = collect_docs_for_audit(folder_token)
    by_type = {}
    for meta in doc_map.values():
        t = meta.get("file_type", "docx")
        by_type[t] = by_type.get(t, 0) + 1
    type_summary = ", ".join(f"{k}={v}" for k, v in sorted(by_type.items()))
    print(f"📄 待审核文件数: {len(doc_map)}（{type_summary}）")

    results = []
    for token, meta in doc_map.items():
        results.append(audit_file(
            token,
            file_type=meta.get("file_type", "docx"),
            source="batch",
            title=meta.get("title"),
        ))

    passed = sum(1 for r in results if r.get("passed") is True)
    failed = sum(1 for r in results if r.get("passed") is False)
    errors = sum(1 for r in results if r.get("passed") is None)
    skipped = sum(
        1 for r in results
        if r.get("passed") is None and r.get("file_type") in UNSUPPORTED_CONTENT_TYPES
    )

    summary = {
        "total": len(results),
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "skipped_unsupported": skipped,
        "by_type": by_type,
        "file_tokens": list(doc_map.keys()),
        "results": results,
    }
    print(
        f"📊 全量审核完成: 共 {len(results)} 个，通过 {passed}，"
        f"未通过 {failed}，失败 {errors}，不支持类型 {skipped}"
    )
    return summary


# ==========================================
# 超管权限审计导出 → 批量审核（生产路径）
# ==========================================
def _normalize_header(name):
    return re.sub(r"\s+", "", (name or "").strip().lower())


def _pick_column(fieldnames, hints):
    normalized = {_normalize_header(n): n for n in fieldnames if n}
    for hint in hints:
        key = _normalize_header(hint)
        if key in normalized:
            return normalized[key]
    for col in fieldnames:
        ncol = _normalize_header(col)
        if any(_normalize_header(h) in ncol for h in hints):
            return col
    return None


def _looks_like_doc_token(value):
    if not value:
        return False
    token = str(value).strip()
    return len(token) >= 20 and re.fullmatch(r"[A-Za-z0-9_-]+", token)


def _read_csv_rows(filepath):
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(filepath, "r", encoding=encoding, newline="") as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法读取 CSV 编码: {filepath}")


def _read_xlsx_rows(filepath):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("读取 xlsx 需要安装 openpyxl: pip install openpyxl") from e

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    result = []
    for row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[i] if i < len(row) else None
        result.append(item)
    return result


def parse_admin_export(filepath):
    """
    解析飞书管理后台「权限审计」导出的 CSV/XLSX。
    返回 [{token, title, doc_type}, ...]
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        rows = _read_csv_rows(filepath)
    elif ext in (".xlsx", ".xls"):
        rows = _read_xlsx_rows(filepath)
    else:
        raise ValueError("仅支持 .csv / .xlsx 文件")

    if not rows:
        return []

    fieldnames = list(rows[0].keys())
    token_col = _pick_column(fieldnames, TOKEN_COLUMN_HINTS)
    title_col = _pick_column(fieldnames, TITLE_COLUMN_HINTS)
    type_col = _pick_column(fieldnames, TYPE_COLUMN_HINTS)

    if not token_col:
        raise ValueError(
            f"未找到文档 token 列。当前列名: {fieldnames}。"
            "请确认导出文件包含「云文档 ID / 文档 token」等列。"
        )

    docs = []
    seen = set()
    for row in rows:
        token = str(row.get(token_col) or "").strip()
        if not _looks_like_doc_token(token) or token in seen:
            continue

        doc_type = str(row.get(type_col) or "docx").strip().lower() if type_col else "docx"
        if type_col and doc_type not in DOCX_TYPE_VALUES:
            continue

        title = str(row.get(title_col) or "").strip() if title_col else None
        docs.append({"token": token, "title": title or None, "doc_type": "docx"})
        seen.add(token)

    return docs


def save_audit_report(summary, prefix="enterprise"):
    os.makedirs(AUDIT_REPORT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(AUDIT_REPORT_DIR, f"{prefix}_{ts}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    return path


def audit_from_admin_export(filepath, doc_types=None):
    """
    从超管权限审计导出文件批量审核。
    doc_types: 默认只审 docx；传 None 表示按导出文件类型列过滤。
    """
    print(f"📥 读取权限审计导出: {filepath}")
    docs = parse_admin_export(filepath)
    print(f"📄 解析到 {len(docs)} 篇待审核 docx 文档")

    if not docs:
        summary = {
            "source_file": filepath,
            "total": 0,
            "passed": 0,
            "failed": 0,
            "errors": 0,
            "permission_errors": 0,
            "results": [],
            "hint": "导出文件为空或未识别到 docx token 列",
        }
        report_path = save_audit_report(summary, prefix="enterprise")
        summary["report_path"] = report_path
        return summary

    results = []
    permission_errors = 0
    for item in docs:
        token = item["token"]
        title = item.get("title")
        register_known_doc(token, title=title, source="admin_export")
        result = audit_document(token, source="enterprise_export", title=title)
        results.append(result)
        err = (result.get("error") or "").lower()
        if result.get("passed") is None and any(k in err for k in ("permission", "forbidden", "权限", "403")):
            permission_errors += 1

    passed = sum(1 for r in results if r.get("passed") is True)
    failed = sum(1 for r in results if r.get("passed") is False)
    errors = sum(1 for r in results if r.get("passed") is None)

    summary = {
        "source_file": filepath,
        "total": len(results),
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "permission_errors": permission_errors,
        "results": results,
    }
    if permission_errors:
        summary["next_step"] = (
            "部分文档应用无读权限。请在飞书文档中为审核应用批量添加「文档应用」，"
            "或由超管通过协作者 API 批量授权后重跑。"
        )

    report_path = save_audit_report(summary, prefix="enterprise")
    summary["report_path"] = report_path
    print(f"📊 企业批量审核完成: 共 {len(results)}，通过 {passed}，未通过 {failed}，失败 {errors}，权限失败 {permission_errors}")
    print(f"📝 报告已保存: {report_path}")
    return summary


# ==========================================
# 用户行为采集（对照采集表）
# ==========================================
def _perm_label(perm):
    if not perm:
        return None
    return PERM_LABELS.get(perm, perm)


def _link_share_label(entity):
    if not entity:
        return None
    return LINK_SHARE_LABELS.get(entity, entity)


def _ts_iso(ts):
    """毫秒/秒时间戳 → ISO 字符串"""
    if ts is None:
        return None
    try:
        val = int(ts)
        if val > 1_000_000_000_000:
            val //= 1000
        return datetime.fromtimestamp(val).isoformat(timespec="seconds")
    except (TypeError, ValueError, OSError):
        return str(ts)


def fetch_file_metadata(file_token, file_type="docx"):
    """拉取文档名称、创建时间、所有者等元数据"""
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    ftype = normalize_file_type(file_type)
    url = "https://open.feishu.cn/open-apis/drive/v1/metas/batch_query"
    payload = {"request_docs": [{"doc_token": file_token, "doc_type": ftype}]}
    try:
        res = requests.post(url, headers=headers, json=payload, timeout=15)
        data = res.json()
    except requests.RequestException as e:
        return None, str(e)

    if data.get("code") != 0:
        return None, data.get("msg", "获取元数据失败")

    metas = data.get("data", {}).get("metas") or []
    if not metas:
        return None, "无元数据"

    meta = metas[0]
    return {
        "title": meta.get("title") or meta.get("name"),
        "create_time": meta.get("create_time"),
        "owner_id": meta.get("owner_id"),
        "latest_modify_time": meta.get("latest_modify_time"),
        "latest_modify_user": meta.get("latest_modify_user"),
        "url": meta.get("url"),
    }, None


def resolve_knowledge_base(file_token, folder_token=None):
    """解析文档所属知识库/文件夹"""
    docs = load_known_docs()
    entry = docs.get(file_token) or {}
    if entry.get("wiki_space_id"):
        return {
            "type": "wiki",
            "space_id": entry.get("wiki_space_id"),
            "space_name": entry.get("wiki_space_name"),
            "node_title": entry.get("title"),
        }
    folder = folder_token or entry.get("folder_token") or AUDIT_ROOT_FOLDER_TOKEN
    if folder:
        return {"type": "folder", "folder_token": folder}
    return None


def get_public_permission(file_token, file_type="docx"):
    """获取链接分享/公开权限设置"""
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    ftype = normalize_file_type(file_type)
    url = f"https://open.feishu.cn/open-apis/drive/v1/permissions/{file_token}/public"
    try:
        res = requests.get(url, headers=headers, params={"type": ftype}, timeout=15)
        data = res.json()
    except requests.RequestException as e:
        return None, str(e)

    if data.get("code") != 0:
        return None, data.get("msg", "获取分享设置失败")

    perm = data.get("data", {}).get("permission_public") or {}
    return {
        "link_share_entity": perm.get("link_share_entity"),
        "link_share_label": _link_share_label(perm.get("link_share_entity")),
        "external_access": perm.get("external_access"),
        "share_entity": perm.get("share_entity"),
        "security_entity": perm.get("security_entity"),
        "comment_entity": perm.get("comment_entity"),
        "invite_external": perm.get("invite_external"),
        "lock_switch": perm.get("lock_switch"),
        "expiration": None,
        "expiration_note": "飞书公开 API 未返回链接有效期，需管理后台审计日志补充",
    }, None


def _extract_rich_text(content_obj):
    """从评论/块富文本结构提取纯文本与 @ 用户（含 text_run / person / docs_link）"""
    if content_obj is None:
        return "", []
    if isinstance(content_obj, str):
        return content_obj, re.findall(r"@(\S+)", content_obj)

    texts = []
    mentioned = []
    if isinstance(content_obj, dict):
        elem_type = content_obj.get("type")
        if elem_type == "text_run":
            tr = content_obj.get("text_run") or {}
            if tr.get("text"):
                texts.append(str(tr["text"]))
        elif elem_type == "person":
            person = content_obj.get("person") or {}
            uid = person.get("user_id") or person.get("open_id") or person.get("union_id")
            if uid:
                mentioned.append(uid)
                texts.append(f"@{uid}")
        elif elem_type == "docs_link":
            link = content_obj.get("docs_link") or {}
            label = link.get("title") or link.get("url") or ""
            if label:
                texts.append(str(label))
        elif elem_type in ("mention", "at_user"):
            uid = (
                content_obj.get("mention_user_id")
                or content_obj.get("user_id")
                or content_obj.get("open_id")
            )
            if uid:
                mentioned.append(uid)
                texts.append(f"@{uid}")

        if content_obj.get("text"):
            texts.append(str(content_obj["text"]))
        for elem in content_obj.get("elements") or []:
            t, m = _extract_rich_text(elem)
            texts.append(t)
            mentioned.extend(m)
        if not elem_type and (content_obj.get("mention_user_id") or content_obj.get("user_id")):
            uid = content_obj.get("mention_user_id") or content_obj.get("user_id") or content_obj.get("open_id")
            if uid:
                mentioned.append(uid)
    elif isinstance(content_obj, list):
        for item in content_obj:
            t, m = _extract_rich_text(item)
            texts.append(t)
            mentioned.extend(m)

    text = "".join(texts).strip()
    return text, list(dict.fromkeys(mentioned))


def _parse_reply_item(reply):
    """解析单条评论回复"""
    text, mentioned = _extract_rich_text(reply.get("content"))
    return {
        "reply_id": reply.get("reply_id") or reply.get("id"),
        "commenter": _user_brief({"user_id": reply.get("user_id")}),
        "content": text,
        "mentioned_users": mentioned,
        "create_time": _ts_iso(reply.get("create_time")),
        "update_time": _ts_iso(reply.get("update_time")),
    }


def fetch_comment_replies(file_token, file_type, comment_id, page_token=None):
    """GET 评论回复列表（局部评论正文通常在此）"""
    headers = _auth_headers()
    if not headers:
        return None, False, None, "无法获取 token"

    ftype = normalize_file_type(file_type)
    params = {"file_type": ftype, "page_size": 50}
    if page_token:
        params["page_token"] = page_token
    url = (
        f"https://open.feishu.cn/open-apis/drive/v1/files/{file_token}"
        f"/comments/{comment_id}/replies"
    )
    try:
        res = requests.get(url, headers=headers, params=params, timeout=15)
        data = res.json()
    except requests.RequestException as e:
        return None, False, None, str(e)

    if data.get("code") != 0:
        return None, False, None, data.get("msg", "获取评论回复失败")

    payload = data.get("data", {})
    return payload.get("items") or [], payload.get("has_more", False), payload.get("page_token"), None


def fetch_all_comment_replies(file_token, file_type, comment_id):
    """拉取某条评论的全部回复"""
    replies = []
    page_token = None
    while True:
        items, has_more, page_token, err = fetch_comment_replies(
            file_token, file_type, comment_id, page_token,
        )
        if err:
            return None, err
        replies.extend(items or [])
        if not has_more:
            break
        if not page_token:
            break
    return replies, None


def fetch_comments_batch(file_token, file_type, comment_ids):
    """batch_query 批量获取评论详情（含局部评论 quote）"""
    if not comment_ids:
        return [], None

    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    ftype = normalize_file_type(file_type)
    url = f"https://open.feishu.cn/open-apis/drive/v1/files/{file_token}/comments/batch_query"
    try:
        res = requests.post(
            url,
            headers=headers,
            params={"file_type": ftype},
            json={"comment_ids": [str(c) for c in comment_ids]},
            timeout=15,
        )
        data = res.json()
    except requests.RequestException as e:
        return None, str(e)

    if data.get("code") != 0:
        return None, data.get("msg", "批量获取评论失败")

    return data.get("data", {}).get("items") or [], None


def _merge_comment_replies(comment, file_token, file_type):
    """列表/batch 结果若缺回复正文，则补拉 replies API"""
    if not comment:
        return comment

    cid = comment.get("comment_id") or comment.get("id")
    existing = comment.get("reply_list", {}).get("replies") or comment.get("replies") or []
    has_content = any(_extract_rich_text(r.get("content"))[0] for r in existing)

    if existing and has_content:
        return comment

    replies, err = fetch_all_comment_replies(file_token, file_type, cid)
    if err or not replies:
        return comment

    merged = dict(comment)
    merged["reply_list"] = {"replies": replies}
    return merged


def _parse_comment_item(comment, file_token=None, file_type=None, reply_id=None):
    """解析评论（含局部评论 quote + 回复正文）"""
    if file_token and file_type:
        comment = _merge_comment_replies(comment, file_token, file_type)

    replies = comment.get("reply_list", {}).get("replies") or comment.get("replies") or []
    quote = (comment.get("quote") or "").strip()
    is_whole = comment.get("is_whole")
    content_parts = []
    mentioned = []
    commenter = _user_brief({"user_id": comment.get("user_id")})
    target_reply_id = str(reply_id) if reply_id else None
    parsed_replies = []

    for reply in replies:
        parsed_reply = _parse_reply_item(reply)
        parsed_replies.append(parsed_reply)
        rid = str(parsed_reply.get("reply_id") or "")
        if target_reply_id and rid != target_reply_id:
            continue
        if parsed_reply.get("content"):
            content_parts.append(parsed_reply["content"])
        mentioned.extend(parsed_reply.get("mentioned_users") or [])
        if parsed_reply.get("commenter"):
            commenter = parsed_reply["commenter"]

    if not content_parts and not target_reply_id:
        text, m = _extract_rich_text(comment.get("content"))
        if text:
            content_parts.append(text)
        mentioned.extend(m)

    content = "\n".join(p for p in content_parts if p)
    if quote and not is_whole:
        content = f"「{quote}」\n{content}".strip() if content else f"「{quote}」"

    return {
        "comment_id": comment.get("comment_id") or comment.get("id"),
        "reply_id": target_reply_id or (parsed_replies[-1].get("reply_id") if parsed_replies else None),
        "commenter": commenter,
        "content": content,
        "quote": quote or None,
        "mentioned_users": list(dict.fromkeys(mentioned)),
        "is_whole": is_whole,
        "replies": parsed_replies,
        "update_time": _ts_iso(comment.get("update_time") or comment.get("create_time")),
    }


def fetch_comment_detail(file_token, file_type, comment_id, reply_id=None):
    """
    获取单条评论完整信息：
    1. batch_query（含 quote）
    2. 补拉 replies API（局部评论正文）
    3. 兜底遍历评论列表
    """
    if not file_token or not comment_id:
        return None, "缺少 file_token 或 comment_id"

    items, err = fetch_comments_batch(file_token, file_type, [comment_id])
    if not err and items:
        for item in items:
            cid = item.get("comment_id") or item.get("id")
            if str(cid) == str(comment_id):
                return _parse_comment_item(
                    item, file_token, file_type, reply_id=reply_id,
                ), None

    page_token = None
    while True:
        list_items, has_more, page_token, err = fetch_file_comments(
            file_token, file_type, page_token,
        )
        if err:
            break
        for item in list_items or []:
            cid = item.get("comment_id") or item.get("id")
            if str(cid) != str(comment_id):
                continue
            return _parse_comment_item(
                item, file_token, file_type, reply_id=reply_id,
            ), None
        if err or not has_more:
            break

    replies, err = fetch_all_comment_replies(file_token, file_type, comment_id)
    if err:
        return None, err
    if not replies:
        return None, "未找到评论"

    stub = {"comment_id": comment_id, "reply_list": {"replies": replies}, "is_whole": False}
    return _parse_comment_item(stub, file_token, file_type, reply_id=reply_id), None


def fetch_comment_by_id(file_token, file_type, comment_id, reply_id=None):
    """兼容旧接口"""
    return fetch_comment_detail(file_token, file_type, comment_id, reply_id=reply_id)


def _comment_record_key(comment, reply_id=None):
    cid = comment.get("comment_id") or comment.get("id")
    if isinstance(comment, dict) and comment.get("reply_id"):
        reply_id = comment.get("reply_id")
    if not reply_id:
        replies = comment.get("reply_list", {}).get("replies") or comment.get("replies") or []
        if replies:
            reply_id = replies[-1].get("reply_id") or replies[-1].get("id")
    if reply_id:
        return f"{cid}:{reply_id}"
    return str(cid)


def _mark_comment_seen(file_token, comment_id, reply_id=None):
    """Webhook 收到评论后写入快照，避免轮询重复记录"""
    if not file_token or not comment_id:
        return
    key = _comment_record_key({"comment_id": comment_id}, reply_id=reply_id)
    snapshot = _load_json_file(COMMENT_SNAPSHOT_FILE) or {"files": {}}
    files = snapshot.setdefault("files", {})
    entry = files.get(file_token) or {"comment_keys": []}
    keys = list(entry.get("comment_keys") or [])
    if key not in keys:
        keys.append(key)
    entry["comment_keys"] = keys
    files[file_token] = entry
    snapshot["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_json_file(COMMENT_SNAPSHOT_FILE, snapshot)


def _extract_block_plain_text(block):
    """从 docx block 提取纯文本"""
    if not isinstance(block, dict):
        return ""
    for key in ("text", "page", "heading1", "heading2", "heading3", "heading4",
                "heading5", "heading6", "heading7", "heading8", "heading9",
                "bullet", "ordered", "quote", "todo", "code"):
        part = block.get(key)
        if isinstance(part, dict):
            text, _ = _extract_rich_text(part)
            if text:
                return text
    return ""


def fetch_docx_block_texts(document_id):
    """分页拉取 docx 全部块文本，用于段落级 diff"""
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    items = []
    page_token = None
    while True:
        params = {"page_size": 500, "document_revision_id": -1}
        if page_token:
            params["page_token"] = page_token
        url = f"https://open.feishu.cn/open-apis/docx/v1/documents/{document_id}/blocks"
        try:
            res = requests.get(url, headers=headers, params=params, timeout=30)
            data = res.json()
        except requests.RequestException as e:
            return None, str(e)

        if data.get("code") != 0:
            return None, data.get("msg", "拉取 blocks 失败")

        payload = data.get("data", {})
        items.extend(payload.get("items") or [])
        if not payload.get("has_more"):
            break
        page_token = payload.get("page_token")
        if not page_token:
            break

    block_map = {}
    for block in items:
        text = _extract_block_plain_text(block)
        if text:
            block_map[block.get("block_id")] = text
    return block_map, None


def _split_paragraphs(text):
    if not text:
        return []
    parts = re.split(r"\n{2,}", text.strip())
    return [p.strip() for p in parts if p.strip()]


def compute_edit_change_summary(file_token, file_type, new_content=None):
    """
    对比上次快照，生成段落级变更摘要。
    docx 优先用 block 级 diff，其它类型用纯文本段落 diff。
    """
    snapshot = _load_json_file(CONTENT_SNAPSHOT_FILE) or {"files": {}}
    old_entry = snapshot.get("files", {}).get(file_token) or {}
    ftype = normalize_file_type(file_type)
    changes = []
    method = "paragraph"

    if ftype == "docx":
        new_blocks, err = fetch_docx_block_texts(file_token)
        if not err and new_blocks is not None:
            method = "block"
            old_blocks = old_entry.get("blocks") or {}
            old_ids = set(old_blocks.keys())
            new_ids = set(new_blocks.keys())
            for bid in sorted(new_ids - old_ids):
                changes.append({
                    "change_type": "added",
                    "block_id": bid,
                    "summary": new_blocks[bid][:200],
                })
            for bid in sorted(old_ids - new_ids):
                changes.append({
                    "change_type": "removed",
                    "block_id": bid,
                    "summary": old_blocks[bid][:200],
                })
            for bid in sorted(old_ids & new_ids):
                if old_blocks[bid] != new_blocks[bid]:
                    changes.append({
                        "change_type": "modified",
                        "block_id": bid,
                        "old_summary": old_blocks[bid][:120],
                        "new_summary": new_blocks[bid][:120],
                    })
            new_content = "\n\n".join(new_blocks.values())

    if not changes:
        if new_content is None:
            new_content, err = fetch_file_content(file_token, file_type)
            if err:
                return {"changes": [], "error": err, "method": method}
        old_content = old_entry.get("content") or ""
        old_paras = _split_paragraphs(old_content)
        new_paras = _split_paragraphs(new_content or "")
        matcher = difflib.SequenceMatcher(None, old_paras, new_paras)
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == "equal":
                continue
            if tag == "insert":
                for idx, para in enumerate(new_paras[j1:j2], start=j1 + 1):
                    changes.append({
                        "change_type": "added",
                        "paragraph_index": idx,
                        "summary": para[:200],
                    })
            elif tag == "delete":
                for idx, para in enumerate(old_paras[i1:i2], start=i1 + 1):
                    changes.append({
                        "change_type": "removed",
                        "paragraph_index": idx,
                        "summary": para[:200],
                    })
            elif tag == "replace":
                changes.append({
                    "change_type": "modified",
                    "paragraph_index": j1 + 1,
                    "old_summary": (old_paras[i1] if i1 < len(old_paras) else "")[:120],
                    "new_summary": (new_paras[j1] if j1 < len(new_paras) else "")[:120],
                })

    save_content_snapshot(file_token, file_type, content=new_content)
    return {
        "method": method,
        "change_count": len(changes),
        "changes": changes[:30],
        "truncated": len(changes) > 30,
    }


def save_content_snapshot(file_token, file_type, content=None):
    """保存文档内容快照，供下次编辑 diff"""
    ftype = normalize_file_type(file_type)
    if content is None:
        content, _ = fetch_file_content(file_token, ftype)

    entry = {
        "file_type": ftype,
        "content": content or "",
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    if ftype == "docx":
        blocks, err = fetch_docx_block_texts(file_token)
        if not err and blocks:
            entry["blocks"] = blocks

    snapshot = _load_json_file(CONTENT_SNAPSHOT_FILE) or {"files": {}}
    snapshot.setdefault("files", {})[file_token] = entry
    _save_json_file(CONTENT_SNAPSHOT_FILE, snapshot)


def fetch_file_comments(file_token, file_type="docx", page_token=None):
    """分页获取文档评论"""
    headers = _auth_headers()
    if not headers:
        return None, False, None, "无法获取 token"

    ftype = normalize_file_type(file_type)
    params = {"file_type": ftype, "page_size": 50}
    if page_token:
        params["page_token"] = page_token
    url = f"https://open.feishu.cn/open-apis/drive/v1/files/{file_token}/comments"
    try:
        res = requests.get(url, headers=headers, params=params, timeout=15)
        data = res.json()
    except requests.RequestException as e:
        return None, False, None, str(e)

    if data.get("code") != 0:
        return None, False, None, data.get("msg", "获取评论失败")

    payload = data.get("data", {})
    return payload.get("items") or [], payload.get("has_more", False), payload.get("page_token"), None


def build_create_collected(file_token, file_type, operator=None, folder_token=None, event_time=None):
    meta, _ = fetch_file_metadata(file_token, file_type)
    known = load_known_docs().get(file_token, {})
    title = (meta or {}).get("title") or known.get("title")
    if title:
        register_known_doc(file_token, title=title, file_type=file_type)
    return {
        "document_name": title,
        "creator": _user_brief(operator) if operator else None,
        "created_at": _ts_iso((meta or {}).get("create_time")) or event_time,
        "knowledge_base": resolve_knowledge_base(file_token, folder_token),
    }


def build_edit_collected(file_token, file_type, operators=None, event_time=None):
    meta, _ = fetch_file_metadata(file_token, file_type)
    editor = None
    if operators:
        editor = _user_brief(operators[0]) if isinstance(operators[0], dict) else operators[0]
    elif meta and meta.get("latest_modify_user"):
        editor = _user_brief(meta.get("latest_modify_user"))

    change_summary = compute_edit_change_summary(file_token, file_type)
    return {
        "editor": editor,
        "edited_at": event_time or _ts_iso((meta or {}).get("latest_modify_time")),
        "change_summary": change_summary,
    }


def build_comment_collected(file_token, file_type, comment_id=None, reply_id=None,
                            from_user=None, is_mentioned=False, event_time=None,
                            notice_type=None, collect_source="api"):
    parsed, err = None, None
    if comment_id and file_token:
        parsed, err = fetch_comment_detail(
            file_token, file_type, comment_id, reply_id=reply_id,
        )
    if parsed:
        mentioned = list(parsed.get("mentioned_users") or [])
        if is_mentioned and from_user:
            uid = (from_user or {}).get("open_id") or (from_user or {}).get("user_id")
            if uid and uid not in mentioned:
                mentioned.append(uid)
        collected = {
            "commenter": parsed.get("commenter") or _user_brief(from_user),
            "content": parsed.get("content"),
            "quote": parsed.get("quote"),
            "reply_id": parsed.get("reply_id") or reply_id,
            "is_whole": parsed.get("is_whole"),
            "mentioned_users": mentioned,
            "commented_at": parsed.get("update_time") or event_time,
            "notice_type": notice_type,
            "collect_source": collect_source,
        }
    else:
        collected = {
            "commenter": _user_brief(from_user),
            "content": None,
            "quote": None,
            "reply_id": reply_id,
            "content_error": err or "需 docs:document.comment:read 权限",
            "mentioned_users": [],
            "commented_at": event_time,
            "notice_type": notice_type,
            "collect_source": collect_source,
        }
    if is_mentioned:
        collected["has_mention_event"] = True
    return collected


def build_forward_collected(operator, member=None, chats=None, link_share_entity=None, event_time=None):
    targets = []
    if member:
        mtype = member.get("member_type") or "user"
        targets.append({
            "target_type": "group" if mtype in ("openchat", "chat") else "user",
            "target_id": member.get("member_id") or member.get("key"),
            "target_name": member.get("name"),
            "permission": _perm_label(member.get("perm")),
        })
    for chat in chats or []:
        if isinstance(chat, dict):
            targets.append({
                "target_type": "group",
                "target_id": chat.get("open_chat_id") or chat.get("chat_id"),
                "target_name": chat.get("name"),
            })
        else:
            targets.append({"target_type": "group", "target_id": chat})

    return {
        "forwarder": _user_brief(operator),
        "forwarded_at": event_time,
        "targets": targets,
        "link_type": _link_share_label(link_share_entity),
    }


def build_share_collected(file_token, file_type, old_perm=None, new_perm=None, event_time=None):
    perm = new_perm or {}
    if not perm and file_token:
        perm, _ = get_public_permission(file_token, file_type)
        perm = perm or {}

    old_link = (old_perm or {}).get("link_share_entity")
    new_link = perm.get("link_share_entity")
    return {
        "link_permission": perm.get("link_share_label") or _link_share_label(new_link),
        "link_share_entity": new_link,
        "external_access": perm.get("external_access"),
        "share_entity": perm.get("share_entity"),
        "expiration": perm.get("expiration"),
        "expiration_note": perm.get("expiration_note"),
        "changed_from": _link_share_label(old_link) if old_link != new_link else None,
        "shared_at": event_time,
    }


def build_delete_collected(operator=None, event_type=None, event_time=None, title=None):
    if event_type == "drive.file.trashed_v1":
        recovery_status = "recycle_bin"
    elif event_type == "drive.file.deleted_v1":
        recovery_status = "permanently_deleted"
    elif event_type == "poll.file_deleted":
        recovery_status = "unknown_removed_or_trashed"
    else:
        recovery_status = "unknown"

    return {
        "deleter": _user_brief(operator),
        "deleted_at": event_time,
        "document_name": title,
        "recovery_status": recovery_status,
        "recovery_status_label": {
            "recycle_bin": "已移入回收站",
            "permanently_deleted": "已彻底删除",
            "unknown_removed_or_trashed": "文件消失(轮询，无法区分回收/移出)",
            "unknown": "未知",
        }.get(recovery_status, recovery_status),
    }


def poll_public_permission_changes(folder_token=None, init_only=False):
    """轮询链接分享设置变更 → 采集表「分享」"""
    doc_map = collect_docs_for_poll(folder_token)
    old_snapshot = _load_json_file(PUBLIC_PERM_SNAPSHOT_FILE)
    is_first_run = not old_snapshot or not old_snapshot.get("files")
    if init_only:
        is_first_run = True

    new_snapshot = {"updated_at": datetime.now().isoformat(timespec="seconds"), "files": {}}
    events = []

    for token, meta in doc_map.items():
        ftype = normalize_file_type(meta.get("file_type", "docx"))
        perm, err = get_public_permission(token, ftype)
        if err:
            if old_snapshot and token in (old_snapshot.get("files") or {}):
                new_snapshot["files"][token] = old_snapshot["files"][token]
            continue

        new_snapshot["files"][token] = {
            "file_type": ftype,
            "title": meta.get("title"),
            "permission": perm,
        }

        if is_first_run:
            continue

        old_perm = (((old_snapshot or {}).get("files") or {}).get(token) or {}).get("permission") or {}
        if old_perm.get("link_share_entity") != perm.get("link_share_entity") or (
            old_perm.get("external_access") != perm.get("external_access")
        ):
            collected = build_share_collected(
                token, ftype, old_perm=old_perm, new_perm=perm,
                event_time=datetime.now().isoformat(timespec="seconds"),
            )
            detail = {"title": meta.get("title"), "old": old_perm, "new": perm, "poll": True}
            record = log_document_event(
                "poll.public_permission_changed", token, ftype, detail,
                source="poll", activity_type="share", collected=collected,
            )
            events.append(record)

    _save_json_file(PUBLIC_PERM_SNAPSHOT_FILE, new_snapshot)
    return {"init_only": is_first_run, "changes": len(events), "events": events}


def poll_comment_changes(folder_token=None, init_only=False):
    """轮询评论列表发现新评论 → 采集表「评论」（webhook 兜底）"""
    doc_map = collect_docs_for_poll(folder_token)
    old_snapshot = _load_json_file(COMMENT_SNAPSHOT_FILE)
    is_first_run = not old_snapshot or not old_snapshot.get("files")
    if init_only:
        is_first_run = True

    new_snapshot = {"updated_at": datetime.now().isoformat(timespec="seconds"), "files": {}}
    events = []

    for token, meta in doc_map.items():
        ftype = normalize_file_type(meta.get("file_type", "docx"))
        if ftype not in AUDITABLE_CONTENT_TYPES:
            continue

        seen_keys = set((((old_snapshot or {}).get("files") or {}).get(token) or {}).get("comment_keys") or [])
        current_keys = []
        page_token = None
        new_comments = []

        while True:
            items, has_more, page_token, err = fetch_file_comments(token, ftype, page_token)
            if err:
                break
            for item in items or []:
                key = _comment_record_key(item)
                current_keys.append(key)
                if not is_first_run and key not in seen_keys:
                    new_comments.append(item)
            if err or not has_more:
                break

        new_snapshot["files"][token] = {
            "file_type": ftype,
            "title": meta.get("title"),
            "comment_keys": current_keys,
        }

        for item in new_comments:
            cid = item.get("comment_id") or item.get("id")
            parsed, err = fetch_comment_detail(token, ftype, cid)
            if not parsed:
                parsed = _parse_comment_item(item, token, ftype)
            collected = {
                "commenter": parsed.get("commenter"),
                "content": parsed.get("content"),
                "quote": parsed.get("quote"),
                "reply_id": parsed.get("reply_id"),
                "is_whole": parsed.get("is_whole"),
                "mentioned_users": parsed.get("mentioned_users"),
                "commented_at": parsed.get("update_time"),
                "collect_source": "poll",
                "content_error": err if not parsed.get("content") else None,
            }
            detail = {"comment": parsed, "poll": True, "title": meta.get("title")}
            record = log_document_event(
                "poll.comment_added", token, ftype, detail,
                source="poll", activity_type="comment", collected=collected,
            )
            events.append(record)

    _save_json_file(COMMENT_SNAPSHOT_FILE, new_snapshot)
    return {"init_only": is_first_run, "changes": len(events), "events": events}


def init_content_snapshots(folder_token=None):
    """为监控中文档建立内容基线（首次 edit diff 用）"""
    doc_map = collect_docs_for_poll(folder_token)
    count = 0
    for token, meta in doc_map.items():
        ftype = normalize_file_type(meta.get("file_type", "docx"))
        if ftype not in AUDITABLE_CONTENT_TYPES:
            continue
        save_content_snapshot(token, ftype)
        count += 1
    print(f"📸 内容快照基线已建立: {count} 个文档")
    return count


def load_recent_activities(limit=50, activity_type=None):
    """按采集表 activity_type 过滤活动记录"""
    events = load_recent_events(limit=EVENTS_LOG_MAX, event_type=None)
    result = []
    for item in events:
        if activity_type and item.get("activity_type") != activity_type:
            continue
        result.append(item)
        if len(result) >= limit:
            break
    return result


# ==========================================
# 文档事件监听与日志
# ==========================================
def _user_brief(user_obj):
    if not user_obj or not isinstance(user_obj, dict):
        return None
    return {
        k: user_obj.get(k)
        for k in ("user_id", "open_id", "union_id")
        if user_obj.get(k)
    }


def _users_brief(user_list):
    if not user_list:
        return []
    return [_user_brief(u) for u in user_list if isinstance(u, dict)]


def log_document_event(event_type, file_token, file_type, event_payload, header=None,
                       audit_result=None, source="webhook", activity_type=None, collected=None):
    """持久化文档事件（webhook / 轮询 / 埋点），并附带采集表字段"""
    record = {
        "event_id": (header or {}).get("event_id") or f"poll-{uuid.uuid4().hex[:16]}",
        "event_type": event_type,
        "event_label": MONITORED_EVENT_TYPES.get(event_type, event_type),
        "activity_type": activity_type,
        "activity_label": ACTIVITY_TYPES.get(activity_type) if activity_type else None,
        "collected": collected or {},
        "source": source,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "create_time": (header or {}).get("create_time"),
        "file_token": file_token,
        "file_type": file_type,
        "detail": event_payload,
    }
    if audit_result is not None:
        record["audit_result"] = {
            "passed": audit_result.get("passed"),
            "hits": audit_result.get("hits"),
            "summary": audit_result.get("summary"),
            "error": audit_result.get("error"),
        }

    try:
        with open(EVENTS_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except OSError as e:
        print(f"⚠️ 事件日志写入失败: {e}")

    _trim_events_log()
    return record


def _trim_events_log():
    if not os.path.exists(EVENTS_LOG_FILE):
        return
    try:
        with open(EVENTS_LOG_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        if len(lines) <= EVENTS_LOG_MAX:
            return
        with open(EVENTS_LOG_FILE, "w", encoding="utf-8") as f:
            f.writelines(lines[-EVENTS_LOG_MAX:])
    except OSError:
        pass


def load_recent_events(limit=50, event_type=None):
    if not os.path.exists(EVENTS_LOG_FILE):
        return []
    try:
        with open(EVENTS_LOG_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except OSError:
        return []

    events = []
    for line in reversed(lines):
        line = line.strip()
        if not line:
            continue
        try:
            item = json.loads(line)
        except json.JSONDecodeError:
            continue
        if event_type and item.get("event_type") != event_type:
            continue
        events.append(item)
        if len(events) >= limit:
            break
    return events


def _permission_event_detail(event_type, event):
    detail = {
        "operator": _user_brief(event.get("operator_id")),
        "added_or_removed_users": _users_brief(event.get("user_list")),
        "added_or_removed_chats": event.get("chat_list") or [],
    }
    if event_type == "drive.file.permission_member_applied_v1":
        detail["permission"] = event.get("permission")
        detail["application_remark"] = event.get("application_remark")
        detail["approver"] = _user_brief(event.get("approver_id"))
    if event_type == "drive.file.created_in_folder_v1":
        detail["folder_token"] = event.get("folder_token")
        detail["operator"] = _user_brief(event.get("operator_id"))
    if event_type == "drive.file.edit_v1":
        detail["operators"] = _users_brief(event.get("operator_id_list"))
    if event_type in ("drive.file.trashed_v1", "drive.file.deleted_v1"):
        detail["operator"] = _user_brief(event.get("operator_id"))
    return detail


def _log_webhook_raw(data):
    """记录所有 webhook 原始请求，便于排查「收不到 POST」"""
    try:
        header = data.get("header", {})
        record = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "event_type": header.get("event_type") or data.get("type"),
            "event_id": header.get("event_id"),
            "schema": data.get("schema"),
            "body": data,
        }
        with open(WEBHOOK_DEBUG_LOG, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except OSError as e:
        print(f"⚠️ webhook 调试日志写入失败: {e}")


# ==========================================
# 方案 B：轮询监听（权限变更 + 文件删除）
# ==========================================
def _load_json_file(path):
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def _save_json_file(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _member_key(member):
    mtype = member.get("member_type") or member.get("type") or "unknown"
    mid = (
        member.get("member_id")
        or member.get("open_id")
        or member.get("member_open_id")
        or member.get("user_id")
        or member.get("id")
        or ""
    )
    return f"{mtype}:{mid}"


def _normalize_member(member):
    return {
        "key": _member_key(member),
        "member_type": member.get("member_type") or member.get("type"),
        "member_id": (
            member.get("member_id")
            or member.get("open_id")
            or member.get("member_open_id")
            or member.get("user_id")
        ),
        "name": member.get("name"),
        "perm": member.get("perm") or member.get("permission"),
    }


def list_permission_members(file_token, file_type):
    """拉取文档协作者列表（分页）"""
    headers = _auth_headers()
    if not headers:
        return None, "无法获取 token"

    ftype = normalize_file_type(file_type)
    if ftype not in POLLABLE_PERMISSION_TYPES:
        return [], None

    members = []
    page_token = None
    while True:
        params = {"type": ftype, "page_size": 100}
        if page_token:
            params["page_token"] = page_token
        url = f"https://open.feishu.cn/open-apis/drive/v1/permissions/{file_token}/members"
        res = requests.get(url, headers=headers, params=params)
        data = res.json()
        if data.get("code") != 0:
            return None, data.get("msg", "获取协作者失败")
        payload = data.get("data", {})
        members.extend(payload.get("items") or [])
        if not payload.get("has_more"):
            break
        page_token = payload.get("page_token")
        if not page_token:
            break
    return members, None


def _members_map(members):
    result = {}
    for m in members or []:
        norm = _normalize_member(m)
        if norm["key"]:
            result[norm["key"]] = norm
    return result


def _compare_members(old_map, new_map):
    changes = []
    for key, member in new_map.items():
        if key not in old_map:
            changes.append(("poll.permission_member_added", member))
        elif old_map[key].get("perm") != member.get("perm"):
            changes.append(("poll.permission_member_changed", {
                **member,
                "old_perm": old_map[key].get("perm"),
                "new_perm": member.get("perm"),
            }))
    for key, member in old_map.items():
        if key not in new_map:
            changes.append(("poll.permission_member_removed", member))
    return changes


def poll_permission_changes(folder_token=None, init_only=False):
    """
    轮询协作者列表，对比快照发现权限变更。
    首次运行或 init_only=True 时只建基线，不产出事件。
    """
    doc_map = collect_docs_for_poll(folder_token)
    old_snapshot = _load_json_file(PERMISSION_SNAPSHOT_FILE)
    is_first_run = not old_snapshot or not old_snapshot.get("files")
    if init_only:
        is_first_run = True

    new_snapshot = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "files": {},
    }
    events = []
    errors = []

    for token, meta in doc_map.items():
        ftype = normalize_file_type(meta.get("file_type", "docx"))
        if ftype not in POLLABLE_PERMISSION_TYPES:
            continue

        members, err = list_permission_members(token, ftype)
        if err:
            errors.append({"file_token": token, "title": meta.get("title"), "error": err})
            if old_snapshot and token in old_snapshot.get("files", {}):
                new_snapshot["files"][token] = old_snapshot["files"][token]
            continue

        member_map = _members_map(members)
        new_snapshot["files"][token] = {
            "file_type": ftype,
            "title": meta.get("title"),
            "members": member_map,
        }

        if is_first_run:
            continue

        old_map = (old_snapshot.get("files", {}).get(token) or {}).get("members") or {}
        for event_type, member in _compare_members(old_map, member_map):
            detail = {
                "member": member,
                "title": meta.get("title"),
                "poll": True,
            }
            label = MONITORED_EVENT_TYPES.get(event_type, event_type)
            print(f"🔍 [轮询-{label}] {token} ({ftype}) {member.get('name') or member.get('key')}")
            activity_type = "forward" if event_type == "poll.permission_member_added" else None
            collected = None
            if activity_type:
                pub, _ = get_public_permission(token, ftype)
                collected = build_forward_collected(
                    operator=None,
                    member=member,
                    link_share_entity=(pub or {}).get("link_share_entity"),
                    event_time=datetime.now().isoformat(timespec="seconds"),
                )
            record = log_document_event(
                event_type, token, ftype, detail, source="poll",
                activity_type=activity_type, collected=collected,
            )
            events.append(record)

    _save_json_file(PERMISSION_SNAPSHOT_FILE, new_snapshot)
    if is_first_run:
        print(f"📸 权限快照已初始化，共 {len(new_snapshot['files'])} 个文件（本次不产生变更事件）")
    else:
        print(f"🔍 权限轮询完成: 变更 {len(events)} 条，失败 {len(errors)} 个文件")

    return {
        "init_only": is_first_run,
        "files_checked": len(new_snapshot["files"]),
        "changes": len(events),
        "errors": errors,
        "events": events,
    }


def _build_folder_files_index(folder_token):
    """构建文件夹内所有文件的 token 索引（含子文件夹）"""
    index = {}
    for f in collect_all_folder_files(folder_token or ""):
        token = f.get("token")
        if not token:
            continue
        index[token] = {
            "name": f.get("name"),
            "type": (f.get("type") or "docx").lower(),
        }
    return index


def poll_file_deletions(folder_token=None, init_only=False):
    """
    轮询共享文件夹文件列表，对比快照发现文件消失（删除/移出）。
    首次运行或 init_only=True 时只建基线。
    """
    folder = folder_token if folder_token is not None else AUDIT_ROOT_FOLDER_TOKEN
    old_snapshot = _load_json_file(FOLDER_FILES_SNAPSHOT_FILE)
    is_first_run = not old_snapshot or not old_snapshot.get("files")
    if init_only:
        is_first_run = True

    current_files = _build_folder_files_index(folder)
    new_snapshot = {
        "folder_token": folder,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "files": current_files,
    }
    events = []

    if not is_first_run:
        old_files = old_snapshot.get("files") or {}
        for token, info in old_files.items():
            if token not in current_files:
                ftype = info.get("type", "docx")
                detail = {
                    "title": info.get("name"),
                    "last_seen_type": ftype,
                    "folder_token": folder,
                    "poll": True,
                }
                print(f"🔍 [轮询-文件已删除] {token} ({ftype}) {info.get('name')}")
                event_time = datetime.now().isoformat(timespec="seconds")
                collected = build_delete_collected(
                    operator=None,
                    event_type="poll.file_deleted",
                    event_time=event_time,
                    title=info.get("name"),
                )
                record = log_document_event(
                    "poll.file_deleted", token, ftype, detail, source="poll",
                    activity_type="delete", collected=collected,
                )
                events.append(record)

    _save_json_file(FOLDER_FILES_SNAPSHOT_FILE, new_snapshot)
    if is_first_run:
        print(f"📸 文件夹快照已初始化，共 {len(current_files)} 个文件（本次不产生删除事件）")
    else:
        print(f"🔍 删除轮询完成: 发现 {len(events)} 个文件消失")

    return {
        "init_only": is_first_run,
        "folder_token": folder,
        "files_count": len(current_files),
        "deleted": len(events),
        "events": events,
    }


def poll_all_monitoring(folder_token=None, init_only=False):
    """执行一轮完整轮询：权限、删除、分享设置、评论"""
    print(f"⏱️  开始轮询监听（间隔配置 {POLL_INTERVAL_SECONDS}s）")
    perm = poll_permission_changes(folder_token, init_only=init_only)
    dele = poll_file_deletions(folder_token, init_only=init_only)
    share = poll_public_permission_changes(folder_token, init_only=init_only)
    comment = poll_comment_changes(folder_token, init_only=init_only)
    summary = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "permission_poll": perm,
        "deletion_poll": dele,
        "share_poll": share,
        "comment_poll": comment,
        "total_changes": (
            perm.get("changes", 0)
            + dele.get("deleted", 0)
            + share.get("changes", 0)
            + comment.get("changes", 0)
        ),
    }
    print(f"✅ 轮询结束，共发现 {summary['total_changes']} 条变更")
    return summary


_poll_thread_started = False


def get_last_webhook_time():
    """最近一次 webhook POST 的时间（用于判断实时链路是否存活）"""
    if not os.path.exists(WEBHOOK_DEBUG_LOG):
        return None
    try:
        with open(WEBHOOK_DEBUG_LOG, "r", encoding="utf-8") as f:
            lines = [ln.strip() for ln in f.readlines() if ln.strip()]
        if not lines:
            return None
        return json.loads(lines[-1]).get("timestamp")
    except (json.JSONDecodeError, OSError):
        return None


def start_background_poller():
    """Flask 启动后在后台按 POLL_INTERVAL_SECONDS 轮询"""
    global _poll_thread_started
    if not AUTO_POLL_ON_START or _poll_thread_started:
        return
    _poll_thread_started = True

    def _loop():
        time.sleep(10)
        while True:
            try:
                poll_all_monitoring()
            except Exception as e:
                print(f"⚠️ 后台轮询异常: {e}")
            time.sleep(POLL_INTERVAL_SECONDS)

    threading.Thread(target=_loop, daemon=True, name="poll-monitor").start()
    print(f"🔄 已启动后台轮询线程，每 {POLL_INTERVAL_SECONDS} 秒检查权限/删除/分享/评论")


# ==========================================
# 飞书事件（功能2：实时审核 + 事件监听）
# ==========================================
def _run_realtime_audit(file_token, file_type, source, title=None):
    ftype = normalize_file_type(file_type)
    if ftype not in AUDITABLE_CONTENT_TYPES and ftype not in UNSUPPORTED_CONTENT_TYPES:
        print(f"ℹ️ 跳过未配置的类型: {file_token} ({file_type})")
        return None
    return audit_file(file_token, file_type=file_type, source=source, title=title)


def handle_feishu_event(event_data):
    """处理飞书 2.0 事件：新建、编辑、权限、删除、评论 + 采集表字段"""
    _log_webhook_raw(event_data)

    header = event_data.get("header", {})
    event_type = header.get("event_type") or event_data.get("type")
    event = event_data.get("event", {})

    if event_type in ("url_verification", None):
        return

    event_time = _ts_iso(header.get("create_time")) or datetime.now().isoformat(timespec="seconds")

    # ---------- 评论事件（drive.notice.comment_add_v1，实时 Webhook）----------
    if event_type == "drive.notice.comment_add_v1":
        meta = event.get("notice_meta") or {}
        file_token = meta.get("file_token")
        file_type = meta.get("file_type", "docx")
        from_user = meta.get("from_user_id")
        comment_id = event.get("comment_id")
        reply_id = event.get("reply_id")
        notice_type = meta.get("notice_type")
        print(f"🔔 [评论-webhook] {file_token} comment_id={comment_id} type={notice_type}")
        if file_token:
            register_known_doc(file_token, source=f"webhook:{event_type}", file_type=file_type)
        collected = build_comment_collected(
            file_token, file_type,
            comment_id=comment_id,
            reply_id=reply_id,
            from_user=from_user,
            is_mentioned=event.get("is_mentioned"),
            event_time=event_time,
            notice_type=notice_type,
            collect_source="webhook",
        )
        detail = {
            "notice_type": notice_type,
            "comment_id": comment_id,
            "reply_id": reply_id,
            "is_mentioned": event.get("is_mentioned"),
            "from_user": _user_brief(from_user),
            "webhook": True,
        }
        record = log_document_event(
            event_type, file_token, file_type, detail, header=header,
            activity_type="comment", collected=collected, source="webhook",
        )
        if file_token and comment_id:
            _mark_comment_seen(file_token, comment_id, reply_id)
        preview = (collected.get("content") or "")[:80]
        print(f"   已记录评论 event_id={record.get('event_id')} content={preview!r}")
        return

    file_token = (
        event.get("file_token")
        or event.get("document_id")
        or header.get("resource_id")
    )
    file_type = event.get("file_type", "docx")

    if event_type not in MONITORED_EVENT_TYPES:
        if event_type and event_type.startswith("drive.file."):
            print(f"📩 收到云文档事件（未列入 MONITORED）: {event_type} file={file_token}")
            log_document_event(event_type, file_token, file_type, event, header=header)
            return
        print(f"📩 忽略非云文档事件: {event_type}")
        return

    label = MONITORED_EVENT_TYPES[event_type]
    print(f"🔔 [{label}] {file_token or '(无token)'} ({file_type})")

    if event_type in MANAGER_RECEIVABLE_EVENTS:
        pass
    elif event_type not in ("drive.file.created_in_folder_v1",):
        print("   ℹ️  权限/删除类事件：需应用为文档所有者 + 开放平台已勾选对应事件")

    if file_token:
        register_known_doc(
            file_token, source=f"webhook:{event_type}", file_type=file_type,
            folder_token=event.get("folder_token"),
        )

    detail = _permission_event_detail(event_type, event)
    audit_result = None
    activity_type = None
    collected = {}

    if event_type == "drive.file.created_in_folder_v1" and file_token:
        activity_type = "create"
        collected = build_create_collected(
            file_token, file_type,
            operator=detail.get("operator"),
            folder_token=detail.get("folder_token"),
            event_time=event_time,
        )
        subscribe_result = auto_subscribe_new_file(
            file_token, file_type, source="auto_new", title=collected.get("document_name"),
        )
        detail["auto_subscribe"] = subscribe_result
        save_content_snapshot(file_token, file_type)
        if event_type in CONTENT_AUDIT_ON_EVENTS:
            audit_result = _run_realtime_audit(
                file_token, file_type, source="realtime_new", title=collected.get("document_name"),
            )

    elif event_type == "drive.file.edit_v1" and file_token:
        activity_type = "edit"
        collected = build_edit_collected(
            file_token, file_type,
            operators=detail.get("operators"),
            event_time=event_time,
        )
        if event_type in CONTENT_AUDIT_ON_EVENTS:
            audit_result = _run_realtime_audit(file_token, file_type, source="realtime_edit")

    elif event_type == "drive.file.permission_member_added_v1":
        activity_type = "forward"
        pub, _ = get_public_permission(file_token, file_type) if file_token else (None, None)
        targets = []
        for u in detail.get("added_or_removed_users") or []:
            if isinstance(u, dict):
                targets.append({
                    "target_type": "user",
                    "target_id": u.get("open_id") or u.get("user_id") or u.get("union_id"),
                })
        for chat in detail.get("added_or_removed_chats") or []:
            if isinstance(chat, dict):
                targets.append({
                    "target_type": "group",
                    "target_id": chat.get("open_chat_id") or chat.get("chat_id"),
                    "target_name": chat.get("name"),
                })
            else:
                targets.append({"target_type": "group", "target_id": chat})
        collected = {
            "forwarder": _user_brief(detail.get("operator")),
            "forwarded_at": event_time,
            "targets": targets,
            "link_type": _link_share_label((pub or {}).get("link_share_entity")),
        }
        perm = detail.get("permission") or ""
        users = detail.get("added_or_removed_users") or []
        print(f"   操作者: {detail.get('operator')}")
        if users:
            print(f"   涉及用户: {len(users)} 人")
        if perm:
            print(f"   申请权限: {perm}")

    elif event_type in (
        "drive.file.permission_member_removed_v1",
        "drive.file.permission_member_applied_v1",
    ):
        perm = detail.get("permission") or ""
        users = detail.get("added_or_removed_users") or []
        print(f"   操作者: {detail.get('operator')}")
        if users:
            print(f"   涉及用户: {len(users)} 人")
        if perm:
            print(f"   申请权限: {perm}")

    elif event_type in ("drive.file.trashed_v1", "drive.file.deleted_v1"):
        activity_type = "delete"
        meta, _ = fetch_file_metadata(file_token, file_type) if file_token else (None, None)
        collected = build_delete_collected(
            operator=detail.get("operator"),
            event_type=event_type,
            event_time=event_time,
            title=(meta or {}).get("title"),
        )
        print(f"   操作者: {detail.get('operator')}")

    record = log_document_event(
        event_type, file_token, file_type, detail, header=header, audit_result=audit_result,
        activity_type=activity_type, collected=collected,
    )
    print(f"   已记录事件 event_id={record.get('event_id')} activity={activity_type}")


def handle_feishu_event_async(event_data):
    threading.Thread(target=handle_feishu_event, args=(event_data,), daemon=True).start()


# ==========================================
# Flask 路由
# ==========================================
@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        data = request.json or {}
        print(f"【根目录收到 POST 请求】: {data}")
        if data.get("type") == "url_verification":
            return jsonify({"challenge": data.get("challenge")})
        handle_feishu_event_async(data)
        return jsonify({"status": "success"})

    return (
        "<h1>Agent 内容审核服务已成功运行！端口: 3000</h1>"
        "<p>接口：</p>"
        "<ul>"
        "<li>POST /audit/all — 全量审核共享文件夹内所有云文档</li>"
        "<li>POST /audit/export — 从权限审计导出文件批量审核</li>"
        "<li>GET /audit/one?file_token=xxx — 审核单篇文档</li>"
        "<li>POST /subscribe/all — 订阅文件夹新建 + 已有文件事件（含权限变更）</li>"
        "<li>GET /events — 查看最近文档事件日志</li>"
        "<li>GET /activities — 按采集表查看用户行为（create/view/edit/…）</li>"
        "<li>GET /activities/schema — 采集表字段说明与覆盖情况</li>"
        "<li>POST /track/view — 埋点：查看（访问者/时长/页数）</li>"
        "<li>POST /track/download — 埋点：下载（导出人/格式）</li>"
        "<li>GET /diagnose — 诊断 subscribe 与事件配置</li>"
        "<li>POST /poll — 立即执行一轮权限/删除/分享/评论轮询</li>"
        "<li>POST /webhook — 飞书实时事件</li>"
        "</ul>"
        "<p>实时(webhook)：新建、编辑、评论(drive.notice.comment_add_v1) | 轮询：权限/转发、删除、分享、评论兜底</p>"
        "<p>评论 Webhook：开放平台添加 drive.notice.comment_add_v1，URL 同 /webhook，权限 docs:document.comment:read</p>"
        "<p>查看/下载：飞书开放 API 无对应事件，请用 /track/view、/track/download 埋点补充</p>"
    )


@app.route("/subscribe", methods=["GET", "POST"])
def subscribe_route():
    if request.method == "POST":
        data = request.json or {}
        file_token = data.get("file_token")
        file_type = data.get("file_type", "docx")
    else:
        file_token = request.args.get("file_token")
        file_type = request.args.get("file_type", "docx")

    if not file_token:
        return jsonify({
            "error": "缺少 file_token",
            "example": "/subscribe?file_token=你的文档token&file_type=docx",
        }), 400

    result = subscribe_document(file_token, file_type)
    if result.get("code") == 0:
        register_known_doc(file_token, source="subscribe", file_type=file_type)
    print(f"【订阅文档】token={file_token}, type={file_type}, 结果={result}")
    return jsonify(result)


@app.route("/subscribe/all", methods=["GET", "POST"])
def subscribe_all_route():
    """订阅共享文件夹新建事件 + 已有文件的全部云文档事件（编辑、权限变更）"""
    folder_token = request.args.get("folder_token") or (request.json or {}).get("folder_token")
    summary = subscribe_all_monitors(folder_token or AUDIT_ROOT_FOLDER_TOKEN)
    return jsonify(summary)


@app.route("/events", methods=["GET"])
def events_route():
    """查看最近文档事件（新建 / 编辑 / 权限变更）"""
    limit = request.args.get("limit", 50, type=int)
    event_type = request.args.get("event_type")
    events = load_recent_events(limit=min(limit, 200), event_type=event_type)
    return jsonify({
        "total": len(events),
        "monitored_event_types": MONITORED_EVENT_TYPES,
        "activity_types": ACTIVITY_TYPES,
        "events": events,
    })


ACTIVITY_SCHEMA = {
    "create": {
        "label": "创建",
        "fields": ["document_name", "creator", "created_at", "knowledge_base"],
        "source": "webhook drive.file.created_in_folder_v1 + 元数据 API",
    },
    "view": {
        "label": "查看",
        "fields": ["visitor", "visited_at", "duration_seconds", "pages_viewed"],
        "source": "POST /track/view 埋点（飞书 API 无查看事件）",
    },
    "edit": {
        "label": "编辑",
        "fields": ["editor", "edited_at", "change_summary"],
        "source": "webhook drive.file.edit_v1 + 内容快照 diff",
    },
    "comment": {
        "label": "评论",
        "fields": ["commenter", "content", "quote", "mentioned_users", "commented_at", "is_whole"],
        "source": "webhook drive.notice.comment_add_v1（实时）+ 轮询；正文经 batch_query/replies API 补全",
    },
    "download": {
        "label": "下载",
        "fields": ["downloader", "downloaded_at", "export_format"],
        "source": "POST /track/download 埋点（飞书 API 无用户下载事件）",
    },
    "forward": {
        "label": "转发",
        "fields": ["forwarder", "targets", "link_type", "forwarded_at"],
        "source": "webhook/轮询 协作者添加（近似转发给人/群）",
    },
    "share": {
        "label": "分享",
        "fields": ["link_permission", "external_access", "expiration", "shared_at"],
        "source": "轮询 GET permissions/public（链接权限；有效期 API 未提供）",
    },
    "delete": {
        "label": "删除",
        "fields": ["deleter", "deleted_at", "document_name", "recovery_status"],
        "source": "webhook trashed/deleted + 轮询文件消失",
    },
}


@app.route("/activities/schema", methods=["GET"])
def activities_schema_route():
    return jsonify({
        "activity_types": ACTIVITY_TYPES,
        "schema": ACTIVITY_SCHEMA,
        "storage": {
            "events_log": EVENTS_LOG_FILE,
            "content_snapshot": CONTENT_SNAPSHOT_FILE,
            "public_permission_snapshot": PUBLIC_PERM_SNAPSHOT_FILE,
            "comment_snapshot": COMMENT_SNAPSHOT_FILE,
        },
    })


@app.route("/activities", methods=["GET"])
def activities_route():
    """按采集表 activity_type 查看用户行为记录"""
    limit = request.args.get("limit", 50, type=int)
    activity_type = request.args.get("activity_type") or request.args.get("type")
    activities = load_recent_activities(limit=min(limit, 200), activity_type=activity_type)
    return jsonify({
        "total": len(activities),
        "activity_type_filter": activity_type,
        "activity_types": ACTIVITY_TYPES,
        "activities": activities,
    })


@app.route("/track/view", methods=["POST"])
def track_view_route():
    """
    埋点：查看行为（飞书开放 API 无法感知用户打开文档）
    body: file_token, file_type, visitor{open_id|user_id}, duration_seconds, pages_viewed
    """
    data = request.json or {}
    file_token = data.get("file_token")
    if not file_token:
        return jsonify({"error": "缺少 file_token"}), 400

    file_type = data.get("file_type", "docx")
    event_time = data.get("visited_at") or datetime.now().isoformat(timespec="seconds")
    collected = {
        "visitor": _user_brief(data.get("visitor")) or data.get("visitor"),
        "visited_at": event_time,
        "duration_seconds": data.get("duration_seconds"),
        "pages_viewed": data.get("pages_viewed"),
        "source_note": data.get("source_note", "client_track"),
    }
    detail = {"track": True, **collected}
    record = log_document_event(
        "track.view", file_token, file_type, detail, source="track",
        activity_type="view", collected=collected,
    )
    return jsonify({"status": "ok", "record": record})


@app.route("/track/download", methods=["POST"])
def track_download_route():
    """
    埋点：下载/导出行为
    body: file_token, file_type, downloader, export_format(pdf|docx|xlsx|csv|...)
    """
    data = request.json or {}
    file_token = data.get("file_token")
    if not file_token:
        return jsonify({"error": "缺少 file_token"}), 400

    file_type = data.get("file_type", "docx")
    event_time = data.get("downloaded_at") or datetime.now().isoformat(timespec="seconds")
    collected = {
        "downloader": _user_brief(data.get("downloader")) or data.get("downloader"),
        "downloaded_at": event_time,
        "export_format": data.get("export_format"),
        "source_note": data.get("source_note", "client_track"),
    }
    detail = {"track": True, **collected}
    record = log_document_event(
        "track.download", file_token, file_type, detail, source="track",
        activity_type="download", collected=collected,
    )
    return jsonify({"status": "ok", "record": record})


@app.route("/comments/fetch", methods=["GET"])
def fetch_comment_route():
    """调试：拉取单条评论完整正文（含局部评论 quote + replies）"""
    file_token = request.args.get("file_token")
    file_type = request.args.get("file_type", "docx")
    comment_id = request.args.get("comment_id")
    reply_id = request.args.get("reply_id")
    if not file_token or not comment_id:
        return jsonify({
            "error": "缺少 file_token 或 comment_id",
            "example": "/comments/fetch?file_token=xxx&comment_id=7657...&file_type=docx",
        }), 400
    parsed, err = fetch_comment_detail(file_token, file_type, comment_id, reply_id=reply_id)
    if err and not parsed:
        return jsonify({"error": err}), 502
    return jsonify({"comment": parsed, "error": err})


@app.route("/snapshot/init", methods=["POST", "GET"])
def snapshot_init_route():
    """初始化内容/权限/评论快照基线（首次 diff / 轮询前调用）"""
    folder_token = request.args.get("folder_token") or (request.json or {}).get("folder_token")
    content_count = init_content_snapshots(folder_token)
    poll_summary = poll_all_monitoring(folder_token, init_only=True)
    return jsonify({
        "content_snapshots": content_count,
        "poll_baselines": poll_summary,
    })


@app.route("/subscribe/folder", methods=["GET", "POST"])
def subscribe_folder_route():
    """订阅文件夹新建事件，用于实时发现新文档"""
    folder_token = request.args.get("folder_token") or (request.json or {}).get("folder_token")
    if not folder_token:
        return jsonify({
            "error": "缺少 folder_token",
            "hint": "在云空间打开目标文件夹，URL 中 /folder/ 后面即为 folder_token",
        }), 400

    result = subscribe_folder(folder_token)
    print(f"【订阅文件夹】token={folder_token}, 结果={result}")
    return jsonify(result)


@app.route("/subscribe/status", methods=["GET"])
def subscribe_status_route():
    file_token = request.args.get("file_token")
    file_type = request.args.get("file_type", "docx")
    if not file_token:
        return jsonify({"error": "缺少 file_token"}), 400
    return jsonify(get_document_subscribe_status(file_token, file_type))


@app.route("/audit/all", methods=["GET", "POST"])
def audit_all_route():
    """功能1：全量审核共享文件夹 / wiki 内所有支持的云文档类型"""
    folder_token = request.args.get("folder_token") or AUDIT_ROOT_FOLDER_TOKEN
    summary = audit_all_documents(folder_token)
    return jsonify(summary)


@app.route("/audit/one", methods=["GET", "POST"])
def audit_one_route():
    file_token = request.args.get("file_token") or (request.json or {}).get("file_token")
    file_type = request.args.get("file_type") or (request.json or {}).get("file_type") or "docx"
    if not file_token:
        return jsonify({"error": "缺少 file_token"}), 400
    result = audit_file(file_token, file_type=file_type, source="api")
    return jsonify(result)


@app.route("/audit/export", methods=["POST"])
def audit_export_route():
    """
    从超管权限审计导出的 CSV/XLSX 批量审核。
    JSON: {"file_path": "E:/exports/audit.xlsx"}
    或 form-data 上传 file 字段。
    """
    upload = request.files.get("file")
    if upload and upload.filename:
        os.makedirs(AUDIT_REPORT_DIR, exist_ok=True)
        tmp_path = os.path.join(AUDIT_REPORT_DIR, f"_upload_{upload.filename}")
        upload.save(tmp_path)
        summary = audit_from_admin_export(tmp_path)
        return jsonify(summary)

    file_path = (request.json or {}).get("file_path") or request.form.get("file_path")
    if not file_path:
        return jsonify({
            "error": "请提供 file_path 或上传 file",
            "example": {"file_path": "E:/exports/permission_audit.xlsx"},
        }), 400
    if not os.path.isfile(file_path):
        return jsonify({"error": f"文件不存在: {file_path}"}), 400

    return jsonify(audit_from_admin_export(file_path))


@app.route("/health", methods=["GET"])
def health_route():
    last_wh = get_last_webhook_time()
    return jsonify({
        "status": "running",
        "realtime": {
            "webhook_path": "/webhook",
            "last_webhook_at": last_wh,
            "comment_event": "drive.notice.comment_add_v1",
            "hint": "编辑/新建/评论实时事件需 Flask + cpolar 同时运行，且飞书事件订阅地址指向当前 cpolar",
        },
        "poll": {
            "auto_poll_on_start": AUTO_POLL_ON_START,
            "interval_seconds": POLL_INTERVAL_SECONDS,
            "note": "权限/删除走轮询，不是 webhook",
        },
    })


@app.route("/diagnose", methods=["GET"])
def diagnose_route():
    return jsonify(diagnose_event_monitoring())


@app.route("/poll", methods=["GET", "POST"])
def poll_route():
    """立即执行轮询（权限/删除/分享/评论）"""
    folder_token = request.args.get("folder_token") or (request.json or {}).get("folder_token")
    init_only = request.args.get("init") == "1" or (request.json or {}).get("init") is True
    summary = poll_all_monitoring(folder_token, init_only=init_only)
    return jsonify(summary)


@app.route("/poll/status", methods=["GET"])
def poll_status_route():
    perm = _load_json_file(PERMISSION_SNAPSHOT_FILE) or {}
    folder = _load_json_file(FOLDER_FILES_SNAPSHOT_FILE) or {}
    public_perm = _load_json_file(PUBLIC_PERM_SNAPSHOT_FILE) or {}
    comments = _load_json_file(COMMENT_SNAPSHOT_FILE) or {}
    content = _load_json_file(CONTENT_SNAPSHOT_FILE) or {}
    return jsonify({
        "poll_interval_seconds": POLL_INTERVAL_SECONDS,
        "auto_poll_on_start": AUTO_POLL_ON_START,
        "permission_snapshot": {
            "updated_at": perm.get("updated_at"),
            "files_count": len(perm.get("files") or {}),
        },
        "folder_snapshot": {
            "updated_at": folder.get("updated_at"),
            "folder_token": folder.get("folder_token"),
            "files_count": len(folder.get("files") or {}),
        },
        "public_permission_snapshot": {
            "updated_at": public_perm.get("updated_at"),
            "files_count": len(public_perm.get("files") or {}),
        },
        "comment_snapshot": {
            "updated_at": comments.get("updated_at"),
            "files_count": len(comments.get("files") or {}),
        },
        "content_snapshot": {
            "files_count": len(content.get("files") or {}),
        },
    })


@app.route("/webhook", methods=["POST"])
def feishu_webhook():
    data = request.json or {}
    event_type = data.get("header", {}).get("event_type", data.get("type"))
    print(f"【/webhook 收到 POST】 event_type={event_type}")

    if data.get("type") == "url_verification":
        challenge = data.get("challenge")
        print(f"【/webhook 挑战验证】challenge={challenge}")
        return jsonify({"challenge": challenge})

    handle_feishu_event_async(data)
    return jsonify({"status": "success"})


# ==========================================
# 命令行入口
# ==========================================
if __name__ == "__main__":
    import sys

    if len(sys.argv) >= 2:
        import json
        cmd = sys.argv[1]

        if cmd == "subscribe" and len(sys.argv) >= 3:
            token = sys.argv[2]
            ftype = sys.argv[3] if len(sys.argv) > 3 else "docx"
            result = subscribe_document(token, ftype)
            if result.get("code") == 0:
                register_known_doc(token, source="subscribe_cli", file_type=ftype)
            print(result)
            sys.exit(0)

        if cmd == "subscribe-folder" and len(sys.argv) >= 3:
            print(subscribe_folder(sys.argv[2]))
            sys.exit(0)

        if cmd == "subscribe-all":
            folder = sys.argv[2] if len(sys.argv) >= 3 else AUDIT_ROOT_FOLDER_TOKEN
            print(json.dumps(subscribe_all_monitors(folder), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "events":
            limit = int(sys.argv[2]) if len(sys.argv) >= 3 else 20
            etype = sys.argv[3] if len(sys.argv) >= 4 else None
            print(json.dumps(load_recent_events(limit=limit, event_type=etype), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "activities":
            limit = int(sys.argv[2]) if len(sys.argv) >= 3 else 20
            atype = sys.argv[3] if len(sys.argv) >= 4 else None
            print(json.dumps(load_recent_activities(limit=limit, activity_type=atype), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "snapshot-init":
            folder = sys.argv[2] if len(sys.argv) >= 3 else None
            print(json.dumps({
                "content_snapshots": init_content_snapshots(folder),
                "poll_baselines": poll_all_monitoring(folder, init_only=True),
            }, ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "fetch-comment" and len(sys.argv) >= 4:
            token, cid = sys.argv[2], sys.argv[3]
            ftype = sys.argv[4] if len(sys.argv) > 4 else "docx"
            rid = sys.argv[5] if len(sys.argv) > 5 else None
            parsed, err = fetch_comment_detail(token, ftype, cid, reply_id=rid)
            print(json.dumps({"comment": parsed, "error": err}, ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "check-subscribe" and len(sys.argv) >= 3:
            ftype = sys.argv[3] if len(sys.argv) > 3 else "docx"
            print(json.dumps(check_subscribe_status(sys.argv[2], ftype), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "diagnose":
            print(json.dumps(diagnose_event_monitoring(), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "poll-init":
            folder = sys.argv[2] if len(sys.argv) >= 3 else AUDIT_ROOT_FOLDER_TOKEN
            print(json.dumps(poll_all_monitoring(folder, init_only=True), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "poll-permissions":
            folder = sys.argv[2] if len(sys.argv) >= 3 else None
            init_only = "--init" in sys.argv
            print(json.dumps(poll_permission_changes(folder, init_only=init_only), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "poll-deletions":
            folder = sys.argv[2] if len(sys.argv) >= 3 else AUDIT_ROOT_FOLDER_TOKEN
            init_only = "--init" in sys.argv
            print(json.dumps(poll_file_deletions(folder, init_only=init_only), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "poll":
            folder = sys.argv[2] if len(sys.argv) >= 3 else None
            print(json.dumps(poll_all_monitoring(folder), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "subscribe-wiki" and len(sys.argv) >= 3:
            print(subscribe_wiki_node(sys.argv[2]))
            sys.exit(0)

        if cmd == "resolve-wiki" and len(sys.argv) >= 3:
            node, err = get_wiki_node(sys.argv[2])
            if err:
                print({"error": err})
            else:
                print(json.dumps(node, ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "audit-wiki" and len(sys.argv) >= 3:
            print(json.dumps(audit_wiki_node(sys.argv[2]), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "audit-export" and len(sys.argv) >= 3:
            summary = audit_from_admin_export(sys.argv[2])
            print(json.dumps(summary, ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "audit-all":
            print(json.dumps(audit_all_documents(), ensure_ascii=False, indent=2))
            sys.exit(0)

        if cmd == "audit" and len(sys.argv) >= 3:
            ftype = sys.argv[3] if len(sys.argv) > 3 else "docx"
            print(json.dumps(audit_file(sys.argv[2], file_type=ftype, source="cli"), ensure_ascii=False, indent=2))
            sys.exit(0)

    ensure_folder_subscribe_on_start()
    start_background_poller()
    print("=" * 60)
    print("✅ 服务已启动 http://0.0.0.0:3000")
    print("📡 实时(webhook)：新建、编辑、评论 → 需 cpolar + 飞书事件订阅 /webhook")
    print(f"🔄 轮询：权限/转发、删除、分享、评论 → 每 {POLL_INTERVAL_SECONDS}s")
    print("📋 用户行为：GET /activities  采集表说明：GET /activities/schema")
    print("📍 埋点补充：POST /track/view  POST /track/download")
    print("💡 首次部署请先：python demo.py snapshot-init")
    print("💡 健康检查: GET /health  查看上次 webhook 时间")
    print("=" * 60)
    # 关闭 reloader，避免后台轮询线程被 debug 模式重启打断
    app.run(host="0.0.0.0", port=3000, debug=True, use_reloader=False)
